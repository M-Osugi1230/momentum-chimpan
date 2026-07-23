"""Build a concise five-to-ten stock daily research plan.

The governed A/B/C/Watch/Skip classification and stored ``daily_action_list``
remain unchanged.  When fewer than five A/B names exist, this presentation layer
adds quality-screened C/Watch rows only to the rendered detailed-research list.
It does not change Momentum score/rank, production priority, or paper execution.
"""
from __future__ import annotations

import html
from pathlib import Path
from typing import Any

import pandas as pd
import yaml
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

POLICY_PATH = "research/daily_research_focus_policy.yaml"
FOCUS_VERSION = "2026-07-23-daily-research-focus-v2"
MINIMUM_DAILY_ACTION_LIST = 5
BUCKET_ORDER = {"A": 0, "B": 1, "C": 2, "Watch": 3, "Skip": 4}


def load_policy(path: str | Path = POLICY_PATH) -> dict[str, Any]:
    payload = yaml.safe_load(Path(path).read_text(encoding="utf-8")) or {}
    if not isinstance(payload, dict):
        raise ValueError("daily research focus policy must be a mapping")
    validate_policy(payload)
    return payload


def validate_policy(payload: dict[str, Any]) -> None:
    policy = payload.get("policy", {})
    limits = payload.get("limits", {})
    boundary = payload.get("governance", {})
    if policy.get("id") != "daily-research-focus-v1":
        raise ValueError("invalid daily research focus policy id")
    if int(limits.get("maximum_A_candidates", 0)) != 5:
        raise ValueError("maximum_A_candidates must be 5")
    if int(limits.get("maximum_daily_action_list", 0)) != 10:
        raise ValueError("maximum_daily_action_list must be 10")
    for key in ("preserve_momentum_score", "preserve_momentum_rank", "preserve_paper_execution"):
        if boundary.get(key) is not True:
            raise ValueError(f"{key} must be true")
    for key in (
        "automatic_score_change",
        "automatic_weight_change",
        "automatic_strategy_change",
        "live_orders",
    ):
        if boundary.get(key) is not False:
            raise ValueError(f"{key} must be false")
    if boundary.get("production_state_mutations") != []:
        raise ValueError("production_state_mutations must be empty")
    if boundary.get("outcome_tracking_required_before_rule_change") is not True:
        raise ValueError("outcome tracking must be required before rule changes")


def optional_text(value: Any) -> str:
    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except (TypeError, ValueError):
        pass
    text = str(value).strip()
    return "" if text.lower() in {"", "nan", "none"} else text


def number(value: Any, default: float | None = None) -> float | None:
    converted = pd.to_numeric(pd.Series([value]), errors="coerce").iloc[0]
    return default if pd.isna(converted) else float(converted)


def normalize_code(value: Any) -> str:
    text = optional_text(value).split(".")[0]
    return text.zfill(4) if text else ""


def unique_parts(parts: list[str]) -> list[str]:
    return list(dict.fromkeys(part.strip() for part in parts if part and part.strip()))


def context_by_code(top100: pd.DataFrame) -> pd.DataFrame:
    columns = [
        "code",
        "rank",
        "score",
        "sector33",
        "is_new_entry",
        "rank_change",
        "is_rising_fast",
        "is_best_rank",
        "top30_streak",
        "top30_streak_days",
        "relative_strength_grade",
        "relative_strength_rank",
        "market_relative_20d",
        "sector_relative_20d",
        "return_5d",
        "return_20d",
        "volume_ratio",
        "trading_value",
        "ma20_deviation",
        "data_quality_grade",
        "data_quality_warnings",
        "data_quality_reason_codes",
        "data_quality_eligible_for_a",
    ]
    if top100 is None or top100.empty or "code" not in top100.columns:
        return pd.DataFrame(columns=columns)
    available = [column for column in columns if column in top100.columns]
    work = top100[available].copy()
    work["code"] = work["code"].map(normalize_code)
    rename = {
        "rank": "ranking_rank_context",
        "score": "ranking_score_context",
    }
    return work.rename(columns=rename).drop_duplicates("code", keep="last")


def base_bucket(row: pd.Series, policy: dict[str, Any]) -> tuple[str, str]:
    priority = optional_text(row.get("action_priority")) or "見送り"
    action_score = number(row.get("action_score"), 0.0) or 0.0
    lifecycle = optional_text(row.get("lifecycle_status"))
    quality = optional_text(row.get("data_quality_grade")) or "D"
    watch = policy["watch_rules"]
    if priority in {"A", "B"}:
        return priority, "既存調査優先度を維持"
    if priority == "C":
        if action_score >= float(watch["C_minimum_action_score"]) and lifecycle in {
            "継続",
            "定着",
            "長期定着",
        }:
            return "C", "継続監視条件を充足"
        return "Watch", "C評価だが継続性またはスコアの改善待ち"
    if (
        action_score >= float(watch["rejected_minimum_action_score"])
        and quality in set(watch["eligible_quality_grades"])
    ):
        return "Watch", "見送りだが閾値に近いため改善待ち"
    return "Skip", "現時点の詳細調査優先度は低い"


def changed_parts(row: pd.Series) -> list[str]:
    parts: list[str] = []
    if bool(row.get("is_new_entry", False)):
        parts.append("Top100新規ランクイン")
    rank_change = number(row.get("rank_change"))
    if bool(row.get("is_rising_fast", False)) and rank_change is not None:
        parts.append(f"前回比+{int(rank_change)}位の急上昇")
    elif rank_change is not None and rank_change > 0:
        parts.append(f"前回比+{int(rank_change)}位")
    elif rank_change is not None and rank_change < 0:
        parts.append(f"前回比{int(rank_change)}位")
    if bool(row.get("is_best_rank", False)):
        parts.append("自己最高順位を更新")
    streak = number(row.get("top30_streak"), number(row.get("top30_streak_days"), 0.0)) or 0.0
    if streak >= 3:
        parts.append(f"Top30を{int(streak)}営業日継続")
    lifecycle = optional_text(row.get("lifecycle_status"))
    if lifecycle:
        parts.append(f"重点候補ライフサイクル: {lifecycle}")
    return unique_parts(parts)


def build_why_today(row: pd.Series) -> str:
    parts = changed_parts(row)
    rank = number(row.get("momentum_rank"), number(row.get("ranking_rank_context")))
    score = number(row.get("momentum_score"), number(row.get("ranking_score_context")))
    if rank is not None:
        parts.append(f"Momentum #{int(rank)}")
    if score is not None:
        parts.append(f"Momentum {int(score)}点")
    relative_grade = optional_text(row.get("relative_strength_grade"))
    relative_rank = number(row.get("relative_strength_rank"))
    if relative_grade:
        detail = f"相対強度{relative_grade}"
        if relative_rank is not None:
            detail += f"・全体{int(relative_rank)}位"
        parts.append(detail)
    sector = optional_text(row.get("sector33"))
    sector_relative = number(row.get("sector_relative_20d"))
    if sector and sector_relative is not None:
        parts.append(f"{sector}中央値比20日{sector_relative:+.1%}")
    positive = optional_text(row.get("positive_reasons"))
    if positive:
        parts.extend(item.strip() for item in positive.split("/") if item.strip())
    return " / ".join(unique_parts(parts)) or "既存の重点候補条件を充足"


def build_change_summary(row: pd.Series) -> str:
    parts = changed_parts(row)
    return " / ".join(parts) if parts else "前回から大きな状態変化はなく継続監視"


def build_risk_summary(row: pd.Series) -> str:
    parts: list[str] = []
    quality = optional_text(row.get("data_quality_grade")) or "D"
    parts.append(f"Data Quality {quality}")
    quality_warning = optional_text(row.get("data_quality_warnings"))
    if quality_warning:
        parts.append(quality_warning)
    caution = optional_text(row.get("caution_reasons"))
    if caution:
        parts.extend(item.strip() for item in caution.split("/") if item.strip())
    liquidity = optional_text(row.get("liquidity_check"))
    if liquidity:
        parts.append(liquidity)
    overheat = optional_text(row.get("overheat_check"))
    if overheat:
        parts.append(overheat)
    return " / ".join(unique_parts(parts)) or "特記事項なし"


def build_research_questions(row: pd.Series, policy: dict[str, Any]) -> str:
    questions = list(policy["research_questions"]["common"])
    if bool(row.get("is_new_entry", False)):
        questions.extend(policy["research_questions"]["new_entry"])
    if bool(row.get("is_rising_fast", False)):
        questions.extend(policy["research_questions"]["rapid_rise"])
    if (number(row.get("volume_ratio"), 0.0) or 0.0) >= 2.0:
        questions.extend(policy["research_questions"]["high_volume"])
    if optional_text(row.get("lifecycle_status")):
        questions.extend(policy["research_questions"]["lifecycle"])
    if optional_text(row.get("data_quality_grade")) in {"B", "C", "D"}:
        questions.extend(policy["research_questions"]["data_warning"])
    return " / ".join(unique_parts(questions)[:6])


def attach_daily_focus(
    action_priority: pd.DataFrame,
    top100: pd.DataFrame,
    policy_path: str | Path = POLICY_PATH,
) -> pd.DataFrame:
    if action_priority is None:
        return pd.DataFrame()
    work = action_priority.copy()
    if work.empty:
        return work
    policy = load_policy(policy_path)
    original_momentum_rank = work["momentum_rank"].copy() if "momentum_rank" in work.columns else None
    original_momentum_score = work["momentum_score"].copy() if "momentum_score" in work.columns else None
    work["code"] = work["code"].map(normalize_code)
    context = context_by_code(top100)
    duplicate_context = [column for column in context.columns if column != "code" and column in work.columns]
    if duplicate_context:
        context = context.drop(columns=duplicate_context)
    work = work.merge(context, on="code", how="left")
    work["daily_focus_version"] = FOCUS_VERSION
    work["action_priority_before_daily_focus"] = work.get(
        "action_priority",
        pd.Series("見送り", index=work.index),
    ).fillna("見送り").astype(str)
    bucket_results = work.apply(lambda row: base_bucket(row, policy), axis=1)
    work["research_bucket"] = [result[0] for result in bucket_results]
    work["focus_adjustment_reason"] = [result[1] for result in bucket_results]

    maximum_a = int(policy["limits"]["maximum_A_candidates"])
    a_rows = work[work["research_bucket"] == "A"].copy()
    if not a_rows.empty:
        sort_columns = []
        ascending = []
        for column, direction in (
            ("action_score", False),
            ("expectancy_score", False),
            ("momentum_rank", True),
        ):
            if column in a_rows.columns:
                sort_columns.append(column)
                ascending.append(direction)
        if sort_columns:
            a_rows = a_rows.sort_values(sort_columns, ascending=ascending)
        excess_indices = a_rows.index[maximum_a:]
        if len(excess_indices):
            work.loc[excess_indices, "research_bucket"] = "B"
            work.loc[excess_indices, "action_priority"] = "B"
            work.loc[excess_indices, "focus_adjustment_reason"] = (
                f"A候補上限{maximum_a}件のためBへ調整"
            )

    work["what_changed"] = work.apply(build_change_summary, axis=1)
    work["why_today"] = work.apply(build_why_today, axis=1)
    work["risk_summary"] = work.apply(build_risk_summary, axis=1)
    work["next_research_questions"] = work.apply(
        lambda row: build_research_questions(row, policy),
        axis=1,
    )
    work["positive_reasons_before_daily_focus"] = work.get(
        "positive_reasons",
        pd.Series("", index=work.index),
    )
    work["caution_reasons_before_daily_focus"] = work.get(
        "caution_reasons",
        pd.Series("", index=work.index),
    )
    work["positive_reasons"] = work["why_today"]
    work["caution_reasons"] = work["risk_summary"]
    work["explanation_complete"] = (
        work["why_today"].astype(str).str.strip().ne("")
        & work["what_changed"].astype(str).str.strip().ne("")
        & work["risk_summary"].astype(str).str.strip().ne("")
        & work["next_research_questions"].astype(str).str.strip().ne("")
    )

    work["_bucket_order"] = work["research_bucket"].map(BUCKET_ORDER).fillna(9)
    sort_columns = ["_bucket_order"]
    ascending = [True]
    for column, direction in (
        ("action_score", False),
        ("expectancy_score", False),
        ("momentum_rank", True),
    ):
        if column in work.columns:
            sort_columns.append(column)
            ascending.append(direction)
    work = work.sort_values(sort_columns, ascending=ascending)
    maximum_actions = int(policy["limits"]["maximum_daily_action_list"])
    actionable = work[work["research_bucket"].isin(["A", "B"])].head(maximum_actions)
    work["daily_action_list"] = work.index.isin(actionable.index)
    action_rank = {index: position for position, index in enumerate(actionable.index, start=1)}
    work["daily_action_rank"] = [action_rank.get(index) for index in work.index]
    work = work.sort_values(
        ["daily_action_list", "daily_action_rank", "_bucket_order", "momentum_rank"],
        ascending=[False, True, True, True],
        na_position="last",
    ).drop(columns="_bucket_order")

    if original_momentum_rank is not None:
        before = original_momentum_rank.reset_index(drop=True)
        after = work.set_index("code").loc[
            action_priority["code"].map(normalize_code), "momentum_rank"
        ].reset_index(drop=True)
        pd.testing.assert_series_equal(after, before, check_names=False)
    if original_momentum_score is not None:
        before = original_momentum_score.reset_index(drop=True)
        after = work.set_index("code").loc[
            action_priority["code"].map(normalize_code), "momentum_score"
        ].reset_index(drop=True)
        pd.testing.assert_series_equal(after, before, check_names=False)
    return work


def _presentation_action_list(focus: pd.DataFrame) -> pd.DataFrame:
    if focus is None or focus.empty:
        return pd.DataFrame(columns=focus.columns if focus is not None else [])
    maximum = 10
    selected = focus[focus.get("daily_action_list", False) == True].sort_values("daily_action_rank").head(maximum).copy()
    selected["daily_action_supplement"] = False
    if len(selected) < MINIMUM_DAILY_ACTION_LIST:
        quality = focus.get("data_quality_grade", pd.Series("D", index=focus.index)).fillna("D").astype(str)
        complete = focus.get("explanation_complete", pd.Series(False, index=focus.index)).fillna(False).astype(bool)
        candidates = focus[
            focus.get("research_bucket", pd.Series("Skip", index=focus.index)).isin(["C", "Watch"])
            & quality.ne("D")
            & complete
            & ~focus.index.isin(selected.index)
        ].copy()
        candidates["_bucket_order"] = candidates["research_bucket"].map(BUCKET_ORDER).fillna(9)
        sort_columns = ["_bucket_order"]
        ascending = [True]
        for column, direction in (("action_score", False), ("expectancy_score", False), ("momentum_rank", True)):
            if column in candidates.columns:
                sort_columns.append(column)
                ascending.append(direction)
        candidates = candidates.sort_values(sort_columns, ascending=ascending).drop(columns="_bucket_order")
        needed = min(MINIMUM_DAILY_ACTION_LIST - len(selected), maximum - len(selected))
        supplements = candidates.head(max(needed, 0)).copy()
        supplements["daily_action_supplement"] = True
        if not supplements.empty:
            selected = pd.concat([selected, supplements], ignore_index=False)
    selected = selected.head(maximum).copy()
    selected["daily_action_rank"] = range(1, len(selected) + 1)
    return selected


def summary_fields(focus: pd.DataFrame) -> dict[str, Any]:
    if focus is None or focus.empty:
        return {
            "Daily Focus A": 0,
            "Daily Focus B": 0,
            "Daily Focus C": 0,
            "Daily Focus Watch": 0,
            "Daily Focus Skip": 0,
            "Daily Action List": 0,
            "Daily Action List補助": 0,
            "Daily Action List下限不足": MINIMUM_DAILY_ACTION_LIST,
            "Daily Focus説明不足": 0,
            "Daily Focus A上限超過": 0,
        }
    buckets = focus.get("research_bucket", pd.Series(index=focus.index, dtype=str))
    selected = _presentation_action_list(focus)
    supplements = selected.get("daily_action_supplement", pd.Series(False, index=selected.index)).fillna(False).astype(bool)
    incomplete = ~focus.get("explanation_complete", pd.Series(False, index=focus.index)).fillna(False).astype(bool)
    return {
        "Daily Focus A": int((buckets == "A").sum()),
        "Daily Focus B": int((buckets == "B").sum()),
        "Daily Focus C": int((buckets == "C").sum()),
        "Daily Focus Watch": int((buckets == "Watch").sum()),
        "Daily Focus Skip": int((buckets == "Skip").sum()),
        "Daily Action List": int(len(selected)),
        "Daily Action List補助": int(supplements.sum()),
        "Daily Action List下限不足": max(MINIMUM_DAILY_ACTION_LIST - len(selected), 0),
        "Daily Focus説明不足": int(incomplete.sum()),
        "Daily Focus A上限超過": max(int((buckets == "A").sum()) - 5, 0),
    }


def action_list(focus: pd.DataFrame) -> pd.DataFrame:
    return _presentation_action_list(focus)


def plain_section(focus: pd.DataFrame) -> list[str]:
    fields = summary_fields(focus)
    lines = [
        "【今日の結論・Daily Action List】",
        "売買推奨ではなく、本日詳しく調査する5〜10社の順番です。",
        f"A {fields['Daily Focus A']}件 / B {fields['Daily Focus B']}件 / C {fields['Daily Focus C']}件 / Watch {fields['Daily Focus Watch']}件 / Skip {fields['Daily Focus Skip']}件",
        f"詳細調査対象 {fields['Daily Action List']}件（目標5〜10件、A最大5件、補助候補{fields['Daily Action List補助']}件）",
    ]
    if fields["Daily Action List下限不足"]:
        lines.append(f"品質条件を満たす候補が少なく、目標下限まであと{fields['Daily Action List下限不足']}件です。無理に追加しません。")
    selected = action_list(focus)
    if selected.empty:
        lines.extend(["本日の詳細調査対象はありません。", ""])
        return lines
    for _, row in selected.iterrows():
        bucket = optional_text(row.get("research_bucket"))
        supplement = "・補助" if bool(row.get("daily_action_supplement", False)) else ""
        lines.extend([
            f"#{int(number(row.get('daily_action_rank'), 0) or 0)} [{bucket}{supplement}] {row.get('code')} {row.get('name')}",
            f"今日の理由：{optional_text(row.get('why_today'))}",
            f"変化：{optional_text(row.get('what_changed'))}",
            f"注意：{optional_text(row.get('risk_summary'))}",
            f"次の確認：{optional_text(row.get('next_research_questions'))}",
            "",
        ])
    return lines


def html_section(focus: pd.DataFrame) -> str:
    fields = summary_fields(focus)
    selected = action_list(focus)
    items = []
    for _, row in selected.iterrows():
        bucket = optional_text(row.get("research_bucket"))
        supplement = bool(row.get("daily_action_supplement", False))
        color = "#166534" if bucket == "A" else "#1d4ed8"
        label = f"{bucket}・補助" if supplement else bucket
        items.append(f'''<div style="border-top:1px solid #e5e7eb;padding:11px 0">
<div style="font-size:14px;font-weight:900;color:#0f172a">#{int(number(row.get("daily_action_rank"), 0) or 0)} [{html.escape(label)}] {html.escape(str(row.get("code", "")))} {html.escape(str(row.get("name", "")))} <span style="float:right;color:{color}">{number(row.get("action_score"), 0):.1f}点</span></div>
<div style="clear:both;font-size:11px;color:{color};font-weight:800;margin-top:4px">今日の理由：{html.escape(optional_text(row.get("why_today")))}</div>
<div style="font-size:11px;color:#475569;margin-top:3px">変化：{html.escape(optional_text(row.get("what_changed")))}</div>
<div style="font-size:11px;color:#b45309;margin-top:3px">注意：{html.escape(optional_text(row.get("risk_summary")))}</div>
<div style="font-size:11px;color:#334155;margin-top:3px">次の確認：{html.escape(optional_text(row.get("next_research_questions")))}</div>
</div>''')
    empty = '<div style="font-size:12px;color:#64748b;margin-top:8px">本日の詳細調査対象はありません。</div>' if not items else ""
    shortfall = ""
    if fields["Daily Action List下限不足"]:
        shortfall = f'<div style="font-size:11px;color:#b45309;margin-top:5px">品質条件を優先し、5件下限まであと{fields["Daily Action List下限不足"]}件は無理に追加していません。</div>'
    return f'''<div style="background:#fff;border:3px solid #0f172a;border-radius:18px;padding:16px;margin-top:14px">
<div style="font-size:20px;font-weight:900;color:#0f172a">今日の結論・Daily Action List</div>
<div style="font-size:12px;color:#64748b;margin-top:4px">売買推奨ではなく、本日詳しく調査する5〜10社の順番です。</div>
<div style="font-size:13px;font-weight:800;color:#334155;margin-top:8px">A {fields["Daily Focus A"]} ・ B {fields["Daily Focus B"]} ・ C {fields["Daily Focus C"]} ・ Watch {fields["Daily Focus Watch"]} ・ Skip {fields["Daily Focus Skip"]}</div>
<div style="font-size:12px;color:#475569;margin-top:4px">詳細調査 {fields["Daily Action List"]}件（目標5〜10件、A最大5件、補助{fields["Daily Action List補助"]}件）</div>{shortfall}{empty}{"".join(items)}</div>'''


def patch_workbook(path: str | Path, focus: pd.DataFrame) -> None:
    target = Path(path)
    if not target.is_file():
        return
    workbook = load_workbook(target)
    if "Daily Action List" in workbook.sheetnames:
        del workbook["Daily Action List"]
    position = 1 if "Summary" in workbook.sheetnames else 0
    sheet = workbook.create_sheet("Daily Action List", position)
    fields = summary_fields(focus)
    summary_rows = [
        ("Policy", "daily-research-focus-v1 / presentation target 5-to-10"),
        ("A", fields["Daily Focus A"]),
        ("B", fields["Daily Focus B"]),
        ("C", fields["Daily Focus C"]),
        ("Watch", fields["Daily Focus Watch"]),
        ("Skip", fields["Daily Focus Skip"]),
        ("Detailed research list", fields["Daily Action List"]),
        ("Supplemental research candidates", fields["Daily Action List補助"]),
        ("Minimum shortfall", fields["Daily Action List下限不足"]),
        ("Incomplete explanations", fields["Daily Focus説明不足"]),
        ("A cap violations", fields["Daily Focus A上限超過"]),
        ("Governed priority-rule mutation", "NONE"),
        ("Score/rank mutation", "NONE"),
        ("Paper execution mutation", "NONE"),
    ]
    sheet.append(["Metric", "Value"])
    for row in summary_rows:
        sheet.append(list(row))
    selected = action_list(focus)
    start_row = len(summary_rows) + 4
    columns = [
        "daily_action_rank",
        "research_bucket",
        "daily_action_supplement",
        "code",
        "name",
        "momentum_rank",
        "momentum_score",
        "action_score",
        "data_quality_grade",
        "what_changed",
        "why_today",
        "risk_summary",
        "next_research_questions",
        "focus_adjustment_reason",
    ]
    available = [column for column in columns if column in selected.columns]
    if selected.empty:
        sheet.cell(start_row, 1, "No detailed research candidates")
    else:
        for column_index, column in enumerate(available, start=1):
            sheet.cell(start_row, column_index, column)
        for row_index, values in enumerate(
            selected[available].itertuples(index=False, name=None), start=start_row + 1
        ):
            for column_index, value in enumerate(values, start=1):
                sheet.cell(row_index, column_index, value)
    header_fill = PatternFill("solid", fgColor="DBEAFE")
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
    if not selected.empty:
        for cell in sheet[start_row]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
    sheet.freeze_panes = f"A{start_row + 1}"
    for column in sheet.columns:
        sheet.column_dimensions[column[0].column_letter].width = min(
            max(len(str(cell.value or "")) for cell in column) + 2, 55
        )
        for cell in column:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    workbook.save(target)
