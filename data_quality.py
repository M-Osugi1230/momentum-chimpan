"""Data-quality grading for daily ranked stocks.

The module adds reliability metadata without changing Momentum score or rank. It
may constrain the human-facing research-priority class, but it never changes
paper execution, production score weights, candidate ranking, or live orders.
"""
from __future__ import annotations

import html
from pathlib import Path
from typing import Any

import pandas as pd
import yaml
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

POLICY_PATH = "research/data_quality_policy.yaml"
QUALITY_VERSION = "2026-07-12-daily-ranking-data-quality-v1"
GRADE_ORDER = {"A": 0, "B": 1, "C": 2, "D": 3}
QUALITY_COLUMNS = [
    "data_quality_version",
    "data_quality_grade",
    "data_quality_score",
    "data_quality_eligible_for_a",
    "data_quality_reason_codes",
    "data_quality_warnings",
    "data_quality_current",
    "data_quality_identity_valid",
    "data_quality_core_complete",
    "data_quality_analytical_complete",
    "data_quality_liquidity_valid",
    "data_quality_corporate_action_suspected",
    "data_quality_abnormal_price",
    "data_quality_abnormal_volume",
]


def load_policy(path: str | Path = POLICY_PATH) -> dict[str, Any]:
    payload = yaml.safe_load(Path(path).read_text(encoding="utf-8")) or {}
    if not isinstance(payload, dict):
        raise ValueError("data quality policy must be a mapping")
    validate_policy(payload)
    return payload


def validate_policy(payload: dict[str, Any]) -> None:
    policy = payload.get("policy", {})
    thresholds = payload.get("thresholds", {})
    grades = payload.get("grades", {})
    boundary = payload.get("priority_boundary", {})
    governance = payload.get("governance", {})
    if policy.get("id") != "daily-ranking-data-quality-v1":
        raise ValueError("invalid data quality policy id")
    if set(grades) != {"A", "B", "C", "D"}:
        raise ValueError("grades must contain exactly A/B/C/D")
    if grades["A"].get("eligible_for_priority_A") is not True:
        raise ValueError("grade A must remain eligible for priority A")
    if grades["B"].get("eligible_for_priority_A") is not True:
        raise ValueError("grade B must remain eligible for priority A")
    if grades["C"].get("eligible_for_priority_A") is not False:
        raise ValueError("grade C must be blocked from priority A")
    if grades["D"].get("eligible_for_priority_A") is not False:
        raise ValueError("grade D must be blocked from priority A")
    if boundary.get("preserve_momentum_score") is not True:
        raise ValueError("Momentum score must be preserved")
    if boundary.get("preserve_momentum_rank") is not True:
        raise ValueError("Momentum rank must be preserved")
    if boundary.get("grade_C_max_priority") != "B":
        raise ValueError("grade C maximum priority must be B")
    if boundary.get("grade_D_forced_priority") != "見送り":
        raise ValueError("grade D must be forced to 見送り")
    for key in (
        "automatic_score_change",
        "automatic_weight_change",
        "automatic_strategy_change",
        "live_orders",
    ):
        if governance.get(key) is not False:
            raise ValueError(f"{key} must be false")
    if governance.get("production_state_mutations") != []:
        raise ValueError("production_state_mutations must be empty")
    for key in (
        "corporate_action_absolute_daily_return",
        "extreme_five_day_absolute_return",
        "extreme_twenty_day_absolute_return",
        "extreme_volume_ratio",
        "extreme_ma20_deviation",
    ):
        value = thresholds.get(key)
        if not isinstance(value, (int, float)) or float(value) <= 0:
            raise ValueError(f"invalid threshold: {key}")


def optional_text(value: Any) -> str:
    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except (TypeError, ValueError):
        pass
    text = str(value).strip()
    return "" if text.lower() in {"", "nan", "none"} else text


def number(value: Any) -> float | None:
    converted = pd.to_numeric(pd.Series([value]), errors="coerce").iloc[0]
    return None if pd.isna(converted) else float(converted)


def normalized_date(value: Any) -> str:
    converted = pd.to_datetime(value, errors="coerce")
    return "" if pd.isna(converted) else converted.date().isoformat()


def append_reason(codes: list[str], warnings: list[str], code: str, warning: str) -> None:
    if code not in codes:
        codes.append(code)
    if warning not in warnings:
        warnings.append(warning)


def evaluate_row(
    row: pd.Series,
    minimum_trading_value: float,
    policy: dict[str, Any],
) -> dict[str, Any]:
    thresholds = policy["thresholds"]
    critical: list[str] = []
    material: list[str] = []
    minor: list[str] = []
    warnings: list[str] = []

    code = optional_text(row.get("code"))
    name = optional_text(row.get("name"))
    expected_date = normalized_date(row.get("date"))
    price_date = normalized_date(row.get("price_date"))
    close = number(row.get("close"))
    volume = number(row.get("volume"))
    trading_value = number(row.get("trading_value"))
    prev_close = number(row.get("prev_close"))
    return_5d = number(row.get("return_5d"))
    return_20d = number(row.get("return_20d"))
    volume_ratio = number(row.get("volume_ratio"))
    ma20 = number(row.get("ma20"))
    ma60 = number(row.get("ma60"))
    ma20_deviation = number(row.get("ma20_deviation"))
    sector = optional_text(row.get("sector33"))

    if not (len(code) == 4 and code.isdigit()):
        append_reason(critical, warnings, "INVALID_CODE", "銘柄コードが4桁の数字ではありません")
    if not name:
        append_reason(critical, warnings, "MISSING_NAME", "銘柄名がありません")
    if not price_date:
        append_reason(critical, warnings, "MISSING_PRICE_DATE", "株価日付がありません")
    if close is None or close <= 0:
        append_reason(critical, warnings, "INVALID_CLOSE", "終値が欠損または不正です")
    if volume is None or volume < 0:
        append_reason(critical, warnings, "INVALID_VOLUME", "出来高が欠損または不正です")
    if trading_value is None or trading_value < 0:
        append_reason(critical, warnings, "INVALID_TRADING_VALUE", "売買代金が欠損または不正です")

    current = bool(expected_date and price_date and expected_date == price_date)
    if expected_date and price_date and not current:
        append_reason(material, warnings, "STALE_PRICE", f"株価日付{price_date}が実行日{expected_date}と一致しません")

    liquidity_valid = bool(trading_value is not None and trading_value >= minimum_trading_value)
    if trading_value is not None and trading_value < minimum_trading_value:
        append_reason(
            material,
            warnings,
            "BELOW_MINIMUM_TRADING_VALUE",
            f"売買代金が基準{minimum_trading_value:,.0f}円未満です",
        )

    missing_analytical = [
        label
        for label, value in (
            ("20日騰落率", return_20d),
            ("出来高倍率", volume_ratio),
            ("20日移動平均", ma20),
            ("60日移動平均", ma60),
        )
        if value is None
    ]
    analytical_complete = not missing_analytical
    if missing_analytical:
        append_reason(
            material,
            warnings,
            "MISSING_ANALYTICAL_FIELD",
            "分析項目不足: " + "、".join(missing_analytical),
        )

    daily_return = None
    if close is not None and prev_close is not None and prev_close > 0:
        daily_return = close / prev_close - 1
    corporate_action = bool(
        daily_return is not None
        and abs(daily_return) >= float(thresholds["corporate_action_absolute_daily_return"])
    )
    if corporate_action:
        append_reason(
            material,
            warnings,
            "POSSIBLE_CORPORATE_ACTION",
            f"前日比{daily_return:.1%}のため株式分割等を要確認です",
        )

    abnormal_price = False
    if return_5d is not None and abs(return_5d) >= float(thresholds["extreme_five_day_absolute_return"]):
        abnormal_price = True
        append_reason(minor, warnings, "EXTREME_5D_RETURN", f"5日騰落率が極端です（{return_5d:.1%}）")
    if return_20d is not None and abs(return_20d) >= float(thresholds["extreme_twenty_day_absolute_return"]):
        abnormal_price = True
        append_reason(minor, warnings, "EXTREME_20D_RETURN", f"20日騰落率が極端です（{return_20d:.1%}）")
    if ma20_deviation is not None and abs(ma20_deviation) >= float(thresholds["extreme_ma20_deviation"]):
        abnormal_price = True
        append_reason(minor, warnings, "EXTREME_MA20_DEVIATION", f"20日線乖離が極端です（{ma20_deviation:.1%}）")

    abnormal_volume = bool(
        volume_ratio is not None
        and volume_ratio >= float(thresholds["extreme_volume_ratio"])
    )
    if abnormal_volume:
        append_reason(minor, warnings, "EXTREME_VOLUME_RATIO", f"出来高倍率が極端です（{volume_ratio:.1f}倍）")
    if not sector:
        append_reason(minor, warnings, "BLANK_SECTOR", "JPX33業種が空欄です")

    if critical:
        grade = "D"
    elif material:
        grade = "C"
    elif minor:
        grade = "B"
    else:
        grade = "A"

    score = max(0, 100 - len(critical) * 100 - len(material) * 25 - len(minor) * 5)
    reason_codes = critical + material + minor
    identity_valid = not any(code in critical for code in ("INVALID_CODE", "MISSING_NAME"))
    core_complete = not any(
        code in critical
        for code in (
            "MISSING_PRICE_DATE",
            "INVALID_CLOSE",
            "INVALID_VOLUME",
            "INVALID_TRADING_VALUE",
        )
    )
    return {
        "data_quality_version": QUALITY_VERSION,
        "data_quality_grade": grade,
        "data_quality_score": int(score),
        "data_quality_eligible_for_a": grade in {"A", "B"},
        "data_quality_reason_codes": "|".join(reason_codes),
        "data_quality_warnings": " / ".join(warnings),
        "data_quality_current": current,
        "data_quality_identity_valid": identity_valid,
        "data_quality_core_complete": core_complete,
        "data_quality_analytical_complete": analytical_complete,
        "data_quality_liquidity_valid": liquidity_valid,
        "data_quality_corporate_action_suspected": corporate_action,
        "data_quality_abnormal_price": abnormal_price,
        "data_quality_abnormal_volume": abnormal_volume,
    }


def attach_quality(
    frame: pd.DataFrame,
    minimum_trading_value: float,
    policy_path: str | Path = POLICY_PATH,
) -> pd.DataFrame:
    if frame is None:
        return pd.DataFrame(columns=QUALITY_COLUMNS)
    work = frame.copy()
    if work.empty:
        for column in QUALITY_COLUMNS:
            if column not in work.columns:
                work[column] = pd.Series(dtype="object")
        return work
    original_score = work["score"].copy() if "score" in work.columns else None
    original_rank = work["rank"].copy() if "rank" in work.columns else None
    policy = load_policy(policy_path)
    quality = work.apply(
        lambda row: pd.Series(evaluate_row(row, minimum_trading_value, policy)),
        axis=1,
    )
    for column in QUALITY_COLUMNS:
        work[column] = quality[column].values
    if original_score is not None:
        pd.testing.assert_series_equal(work["score"], original_score, check_names=False)
    if original_rank is not None:
        pd.testing.assert_series_equal(work["rank"], original_rank, check_names=False)
    return work


def quality_by_code(top100: pd.DataFrame) -> pd.DataFrame:
    columns = ["code", *QUALITY_COLUMNS]
    if top100 is None or top100.empty or "code" not in top100.columns:
        return pd.DataFrame(columns=columns)
    available = [column for column in columns if column in top100.columns]
    result = top100[available].copy()
    result["code"] = result["code"].astype(str).str.split(".").str[0].str.zfill(4)
    return result.drop_duplicates("code", keep="last")


def apply_priority_gate(action_priority: pd.DataFrame, top100: pd.DataFrame) -> pd.DataFrame:
    if action_priority is None:
        return pd.DataFrame()
    work = action_priority.copy()
    if work.empty:
        return work
    if "code" not in work.columns:
        return work
    work["code"] = work["code"].astype(str).str.split(".").str[0].str.zfill(4)
    quality = quality_by_code(top100)
    drop_existing = [column for column in QUALITY_COLUMNS if column in work.columns]
    if drop_existing:
        work = work.drop(columns=drop_existing)
    work = work.merge(quality, on="code", how="left")
    work["data_quality_grade"] = work.get("data_quality_grade", pd.Series(index=work.index, dtype=str)).fillna("D")
    work["data_quality_eligible_for_a"] = work.get(
        "data_quality_eligible_for_a",
        pd.Series(False, index=work.index),
    ).fillna(False).astype(bool)
    original = work.get("action_priority", pd.Series("見送り", index=work.index)).fillna("見送り").astype(str)
    work["action_priority_before_quality"] = original
    adjusted: list[str] = []
    reasons: list[str] = []
    for position, row in work.iterrows():
        grade = optional_text(row.get("data_quality_grade")) or "D"
        current = optional_text(row.get("action_priority_before_quality")) or "見送り"
        if grade == "D":
            new_value = "見送り"
            reason = "品質Dのため調査優先候補から除外"
        elif grade == "C" and current == "A":
            new_value = "B"
            reason = "品質CのためA昇格を禁止しBへ制限"
        else:
            new_value = current
            reason = "品質ゲートによる変更なし"
        adjusted.append(new_value)
        reasons.append(reason)
    work["action_priority"] = adjusted
    work["quality_adjustment_reason"] = reasons
    order = {"A": 0, "B": 1, "C": 2, "見送り": 3}
    work["_quality_priority_order"] = work["action_priority"].map(order).fillna(9)
    sort_columns = ["_quality_priority_order"]
    ascending = [True]
    for column, direction in (
        ("action_priority_score", False),
        ("expectancy_score", False),
        ("rank", True),
    ):
        if column in work.columns:
            sort_columns.append(column)
            ascending.append(direction)
    return work.sort_values(sort_columns, ascending=ascending).drop(columns="_quality_priority_order")


def replace_frame_in_place(target: pd.DataFrame, replacement: pd.DataFrame) -> None:
    if target is replacement:
        return
    target.drop(target.index, inplace=True)
    for column in list(target.columns):
        if column not in replacement.columns:
            target.drop(columns=[column], inplace=True)
    for column in replacement.columns:
        target[column] = replacement[column].reset_index(drop=True)
    target.index = replacement.index


def summary_fields(top100: pd.DataFrame, action_priority: pd.DataFrame | None = None) -> dict[str, Any]:
    frame = top100 if top100 is not None else pd.DataFrame()
    grades = frame.get("data_quality_grade", pd.Series(dtype=str))
    fields: dict[str, Any] = {
        "Data Quality評価件数": int(grades.notna().sum()),
        "Data Quality A": int((grades == "A").sum()),
        "Data Quality B": int((grades == "B").sum()),
        "Data Quality C": int((grades == "C").sum()),
        "Data Quality D": int((grades == "D").sum()),
        "Data Quality A適格率": float(frame.get("data_quality_eligible_for_a", pd.Series(dtype=bool)).fillna(False).mean()) if len(frame) else 0.0,
        "Data Quality現行日率": float(frame.get("data_quality_current", pd.Series(dtype=bool)).fillna(False).mean()) if len(frame) else 0.0,
        "Data Quality corporate action警告": int(frame.get("data_quality_corporate_action_suspected", pd.Series(dtype=bool)).fillna(False).sum()),
    }
    if action_priority is not None and not action_priority.empty:
        before = action_priority.get("action_priority_before_quality", pd.Series(dtype=str))
        after = action_priority.get("action_priority", pd.Series(dtype=str))
        fields["Data Quality優先度調整数"] = int((before.astype(str) != after.astype(str)).sum()) if len(before) == len(after) else 0
        fields["品質C/DのA候補"] = int(
            (
                after.eq("A")
                & action_priority.get("data_quality_grade", pd.Series(index=action_priority.index, dtype=str)).isin(["C", "D"])
            ).sum()
        )
    else:
        fields["Data Quality優先度調整数"] = 0
        fields["品質C/DのA候補"] = 0
    return fields


def quality_table(top100: pd.DataFrame) -> pd.DataFrame:
    columns = [
        "rank",
        "code",
        "name",
        "sector33",
        "score",
        "price_date",
        "data_quality_grade",
        "data_quality_score",
        "data_quality_eligible_for_a",
        "data_quality_current",
        "data_quality_identity_valid",
        "data_quality_core_complete",
        "data_quality_analytical_complete",
        "data_quality_liquidity_valid",
        "data_quality_corporate_action_suspected",
        "data_quality_abnormal_price",
        "data_quality_abnormal_volume",
        "data_quality_reason_codes",
        "data_quality_warnings",
    ]
    if top100 is None or top100.empty:
        return pd.DataFrame(columns=columns)
    available = [column for column in columns if column in top100.columns]
    result = top100[available].copy()
    result["_grade_order"] = result.get("data_quality_grade", pd.Series(index=result.index, dtype=str)).map(GRADE_ORDER).fillna(9)
    sort_columns = ["_grade_order"]
    ascending = [True]
    if "rank" in result.columns:
        sort_columns.append("rank")
        ascending.append(True)
    return result.sort_values(sort_columns, ascending=ascending).drop(columns="_grade_order")


def plain_section(top100: pd.DataFrame, action_priority: pd.DataFrame | None = None) -> list[str]:
    fields = summary_fields(top100, action_priority)
    lines = [
        "【Data Quality】",
        (
            f"Top100: A {fields['Data Quality A']} / B {fields['Data Quality B']} / "
            f"C {fields['Data Quality C']} / D {fields['Data Quality D']}"
        ),
        (
            f"当日データ率 {fields['Data Quality現行日率']:.1%} / "
            f"A適格率 {fields['Data Quality A適格率']:.1%} / "
            f"優先度調整 {fields['Data Quality優先度調整数']}件"
        ),
        "品質CはA昇格不可、品質Dは見送りです。Momentumスコアと順位は変更しません。",
    ]
    cautions = quality_table(top100)
    cautions = cautions[cautions.get("data_quality_grade", pd.Series(dtype=str)).isin(["C", "D"])]
    for _, row in cautions.head(5).iterrows():
        lines.append(
            f"  {row.get('data_quality_grade')} #{int(row.get('rank', 0) or 0)} "
            f"{row.get('code')} {row.get('name')}｜{row.get('data_quality_warnings', '')}"
        )
    lines.append("")
    return lines


def html_section(top100: pd.DataFrame, action_priority: pd.DataFrame | None = None) -> str:
    fields = summary_fields(top100, action_priority)
    cautions = quality_table(top100)
    cautions = cautions[cautions.get("data_quality_grade", pd.Series(dtype=str)).isin(["C", "D"])]
    warning_rows = "".join(
        (
            '<div style="border-top:1px solid #e5e7eb;padding:7px 0;font-size:11px;color:#475569">'
            f'<b>{html.escape(str(row.get("data_quality_grade", "")))} '
            f'#{int(row.get("rank", 0) or 0)} {html.escape(str(row.get("code", "")))} '
            f'{html.escape(str(row.get("name", "")))}</b> ・ '
            f'{html.escape(str(row.get("data_quality_warnings", "")))}</div>'
        )
        for _, row in cautions.head(5).iterrows()
    )
    return f'''<div style="background:#fff;border:2px solid #0f766e;border-radius:18px;padding:16px;margin-top:14px">
<div style="font-size:18px;font-weight:900;color:#115e59">Data Quality</div>
<div style="font-size:13px;color:#334155;margin-top:6px">Top100: A <b>{fields["Data Quality A"]}</b> ・ B <b>{fields["Data Quality B"]}</b> ・ C <b>{fields["Data Quality C"]}</b> ・ D <b>{fields["Data Quality D"]}</b></div>
<div style="font-size:12px;color:#475569;margin-top:5px">当日データ率 <b>{fields["Data Quality現行日率"]:.1%}</b> ・ A適格率 <b>{fields["Data Quality A適格率"]:.1%}</b> ・ 優先度調整 <b>{fields["Data Quality優先度調整数"]}件</b></div>
<div style="font-size:11px;color:#64748b;margin-top:5px">品質CはA昇格不可、品質Dは見送り。Momentumスコアと順位は不変です。</div>{warning_rows}</div>'''


def patch_workbook(path: str | Path, top100: pd.DataFrame, action_priority: pd.DataFrame | None = None) -> None:
    target = Path(path)
    if not target.is_file():
        return
    workbook = load_workbook(target)
    if "Data Quality" in workbook.sheetnames:
        del workbook["Data Quality"]
    position = 1 if "Summary" in workbook.sheetnames else 0
    sheet = workbook.create_sheet("Data Quality", position)
    fields = summary_fields(top100, action_priority)
    summary_rows = [
        ("Policy", "daily-ranking-data-quality-v1"),
        ("Top100 assessed", fields["Data Quality評価件数"]),
        ("Grade A", fields["Data Quality A"]),
        ("Grade B", fields["Data Quality B"]),
        ("Grade C", fields["Data Quality C"]),
        ("Grade D", fields["Data Quality D"]),
        ("Current-date ratio", fields["Data Quality現行日率"]),
        ("Eligible-for-A ratio", fields["Data Quality A適格率"]),
        ("Priority adjustments", fields["Data Quality優先度調整数"]),
        ("C/D remaining in A", fields["品質C/DのA候補"]),
        ("Score/rank mutation", "NONE"),
        ("Automatic strategy/weight change", "DISABLED"),
    ]
    sheet.append(["Metric", "Value"])
    for row in summary_rows:
        sheet.append(list(row))
    start_row = len(summary_rows) + 4
    table = quality_table(top100)
    if table.empty:
        sheet.cell(start_row, 1, "No Top100 rows")
    else:
        for column_index, column in enumerate(table.columns, start=1):
            sheet.cell(start_row, column_index, column)
        for row_index, values in enumerate(table.itertuples(index=False, name=None), start=start_row + 1):
            for column_index, value in enumerate(values, start=1):
                sheet.cell(row_index, column_index, value)
    header_fill = PatternFill("solid", fgColor="D1FAE5")
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
    if not table.empty:
        for cell in sheet[start_row]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
    sheet.freeze_panes = f"A{start_row + 1}"
    for column in sheet.columns:
        width = min(max(len(str(cell.value or "")) for cell in column) + 2, 48)
        sheet.column_dimensions[column[0].column_letter].width = width
        for cell in column:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    workbook.save(target)
