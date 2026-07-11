from __future__ import annotations

import importlib.util
import inspect
import sys
import tempfile
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook


spec = importlib.util.spec_from_file_location("momentum_main", "main.py")
module = importlib.util.module_from_spec(spec)
sys.modules[spec.name] = module
assert spec.loader is not None
spec.loader.exec_module(module)

assert module.APP_VERSION == "2026-07-11-dashboard-relative-strength-v18"
assert module.EXECUTION_MODE == "RESEARCH_AND_PAPER_ONLY"

# Cross-sectional relative strength: the strongest stock must rank above the weakest.
frame = pd.DataFrame([
    {"rank": 1, "code": "1001", "name": "A1", "sector33": "電気機器", "score": 80, "return_20d": 0.30, "return_60d": 0.55, "trading_value": 5e9, "volume_ratio": 2.0},
    {"rank": 2, "code": "1002", "name": "A2", "sector33": "電気機器", "score": 75, "return_20d": 0.10, "return_60d": 0.20, "trading_value": 3e9, "volume_ratio": 1.8},
    {"rank": 3, "code": "1003", "name": "A3", "sector33": "電気機器", "score": 65, "return_20d": -0.05, "return_60d": 0.00, "trading_value": 1e9, "volume_ratio": 1.2},
    {"rank": 4, "code": "2001", "name": "B1", "sector33": "機械", "score": 78, "return_20d": 0.22, "return_60d": 0.40, "trading_value": 4e9, "volume_ratio": 2.2},
    {"rank": 5, "code": "2002", "name": "B2", "sector33": "機械", "score": 62, "return_20d": 0.02, "return_60d": 0.10, "trading_value": 8e8, "volume_ratio": 1.1},
    {"rank": 6, "code": "2003", "name": "B3", "sector33": "機械", "score": 50, "return_20d": -0.12, "return_60d": -0.18, "trading_value": 5e8, "volume_ratio": 0.8},
])
relative = module.attach_relative_strength(frame)
strong = relative.set_index("code").loc["1001"]
weak = relative.set_index("code").loc["2003"]
assert strong["relative_strength_score"] > weak["relative_strength_score"]
assert int(strong["relative_strength_rank"]) == 1
assert bool(strong["dual_outperformer"])
assert strong["market_relative_20d"] > 0
assert strong["sector_relative_20d"] > 0
assert weak["relative_strength_grade"] == "C"

relative_table = module.build_relative_strength_table(relative)
assert relative_table.iloc[0]["code"] == "1001"
assert "市場・業種相対強度" in "\n".join(module.plain_relative_strength_section(relative_table))
assert "市場・業種相対強度" in module.html_relative_strength_section(relative_table)

# Relative strength must influence sector leader score without changing the base Momentum score.
base = pd.Series({
    "score": 75, "rank": 20, "sector_momentum_score": 65, "sector_rotation": "主導",
    "trading_value": 2_000_000_000, "volume_ratio": 2.0, "return_20d": 0.15,
    "ma20_deviation": 0.10, "action_priority": "B", "expectancy_score": 65,
    "expectancy_confidence": "中", "relative_strength_grade": "A",
})
high = base.copy()
high["relative_strength_score"] = 85
low = base.copy()
low["relative_strength_score"] = 25
assert module.sector_leader_values(high)["sector_leader_score"] > module.sector_leader_values(low)["sector_leader_score"]

# Benchmark-adjusted forward results: stock outperforms both market and sector medians.
dates = ["2026-01-05", "2026-01-06"]
price_rows = []
for i in range(12):
    code = f"{3000 + i}"
    sector = "電気機器" if i < 4 else "機械"
    entry = 100.0
    if i == 0:
        exit_price = 110.0
    elif i < 4:
        exit_price = 104.0
    else:
        exit_price = 102.0
    price_rows.extend([
        {"date": dates[0], "code": code, "sector33": sector, "close": entry},
        {"date": dates[1], "code": code, "sector33": sector, "close": exit_price},
    ])
price_history = pd.DataFrame(price_rows)
signal = pd.DataFrame([{
    "signal_date": dates[0], "entry_price_date": dates[0], "code": "3000", "name": "Leader",
    "sector33": "電気機器", "entry_close": 100.0, "sector_research_priority": "最優先",
    "sector_leader_grade": "S", "sector_rotation": "主導", "sector_leader_score": 90.0,
}])
outcomes = module.calculate_sector_leader_outcomes(signal, price_history, horizons=(1,))
assert len(outcomes) == 1
outcome = outcomes.iloc[0]
assert abs(float(outcome["forward_return"]) - 0.10) < 1e-9
assert abs(float(outcome["market_benchmark_return"]) - 0.02) < 1e-9
assert abs(float(outcome["sector_benchmark_return"]) - 0.04) < 1e-9
assert abs(float(outcome["market_excess_return"]) - 0.08) < 1e-9
assert abs(float(outcome["sector_excess_return"]) - 0.06) < 1e-9
assert bool(outcome["market_outperformance"])
assert bool(outcome["sector_outperformance"])
assert int(outcome["market_peer_count"]) == 12
assert int(outcome["sector_peer_count"]) == 4

performance = module.build_sector_leader_performance_summary(outcomes)
overall = performance.iloc[0]
assert abs(float(overall["average_market_excess_return"]) - 0.08) < 1e-9
assert abs(float(overall["average_sector_excess_return"]) - 0.06) < 1e-9
assert float(overall["market_outperformance_rate"]) == 1.0
assert float(overall["sector_outperformance_rate"]) == 1.0

# Excel must include the dedicated relative-strength sheet.
with tempfile.TemporaryDirectory() as tmp:
    output = str(Path(tmp) / "report.xlsx")
    kwargs = {}
    for name in inspect.signature(module.excel_report).parameters:
        if name == "path":
            kwargs[name] = output
        elif name == "summary":
            kwargs[name] = {"test": "ok"}
        elif name == "top100":
            kwargs[name] = relative
        elif name == "relative_strength":
            kwargs[name] = relative_table
        elif name == "errors":
            kwargs[name] = []
        else:
            kwargs[name] = pd.DataFrame()
    module.excel_report(**kwargs)
    workbook = load_workbook(output, read_only=True)
    assert "Relative Strength" in workbook.sheetnames
    assert "Sector Leader Outcomes" in workbook.sheetnames

source = Path("main.py").read_text(encoding="utf-8")
assert "all_ranked = attach_relative_strength(all_ranked)" in source
assert "relative_strength = build_relative_strength_table(top100)" in source
assert "html_relative_strength_section(relative_strength)" in source
assert "market_excess_return" in source
assert "sector_excess_return" in source
assert 'sheet_name="Relative Strength"' in source

print("relative strength and benchmark-adjusted performance validation passed")
