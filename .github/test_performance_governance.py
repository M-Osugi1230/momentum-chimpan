from pathlib import Path
from tempfile import TemporaryDirectory
import importlib.util

import pandas as pd
from openpyxl import load_workbook

spec = importlib.util.spec_from_file_location("momentum_main", "main.py")
module = importlib.util.module_from_spec(spec)
assert spec.loader is not None
spec.loader.exec_module(module)

assert module.APP_VERSION == "2026-07-11-dashboard-performance-governance-v14"

leader_rows = pd.DataFrame([
    {
        "overall_leader_rank": 1,
        "sector_leader_rank": 1,
        "sector33": "電気機器",
        "sector_rank": 1,
        "sector_momentum_score": 82.0,
        "sector_strength": "強い",
        "sector_rotation": "加速",
        "sector_score_delta": 6.0,
        "code": "0001",
        "name": "Alpha",
        "close": 100.0,
        "price_date": "2026-01-05",
        "momentum_rank": 2,
        "momentum_score": 90.0,
        "sector_leader_score": 91.0,
        "sector_leader_grade": "S",
        "sector_research_priority": "最優先",
        "action_priority": "A",
        "action_score": 92.0,
        "expectancy_score": 78.0,
        "expectancy_confidence": "高",
        "return_20d": 0.20,
        "return_60d": 0.35,
        "volume_ratio": 3.2,
        "trading_value": 5_000_000_000,
        "ma20_deviation": 0.10,
        "leader_reasons": "業種加速",
        "leader_cautions": "",
    }
])

snapshot = module.current_sector_signal_snapshot("2026-01-05", leader_rows)
assert len(snapshot) == 1
assert snapshot.iloc[0]["entry_close"] == 100.0
assert snapshot.iloc[0]["sector_research_priority"] == "最優先"

with TemporaryDirectory() as tmpdir:
    history_path = str(Path(tmpdir) / "sector_history.csv")
    first = module.update_sector_signal_history(history_path, snapshot)
    second = module.update_sector_signal_history(history_path, snapshot)
    assert len(first) == 1
    assert len(second) == 1
    assert Path(history_path).exists()

price_dates = pd.bdate_range("2026-01-05", periods=26)
price_history = pd.DataFrame({
    "date": [d.date().isoformat() for d in price_dates],
    "code": ["0001"] * len(price_dates),
    "close": [100.0 + i * 2.0 for i in range(len(price_dates))],
})
outcomes = module.calculate_sector_leader_outcomes(snapshot, price_history)
assert set(outcomes["horizon_days"]) == {5, 10, 20}
assert (outcomes["forward_return"] > 0).all()
assert outcomes["win"].all()

performance = module.build_sector_leader_performance_summary(outcomes)
assert module.performance_overall_stats(performance, 10)["count"] == 1
assert "rotation" in set(performance["group_type"])

# Build a deterministic degradation sample: strong baseline followed by weak recent results.
degradation_rows = []
for index in range(40):
    recent = index >= 20
    degradation_rows.append({
        "signal_date": (pd.Timestamp("2025-01-01") + pd.Timedelta(days=index)).date().isoformat(),
        "entry_price_date": "2025-01-01",
        "exit_price_date": "2025-02-01",
        "code": f"{index:04d}",
        "name": f"Stock{index}",
        "sector33": "電気機器",
        "sector_research_priority": "最優先",
        "sector_leader_grade": "S",
        "sector_rotation": "加速",
        "sector_leader_score": 90.0,
        "horizon_days": 10,
        "entry_close": 100.0,
        "exit_close": 96.0 if recent else 106.0,
        "forward_return": -0.04 if recent else 0.06,
        "win": not recent,
        "calendar_days": 14,
    })
degradation_outcomes = pd.DataFrame(degradation_rows)
governance = module.build_signal_governance(degradation_outcomes)
overall_10 = governance[(governance["scope_type"] == "overall") & (governance["horizon_days"] == 10)].iloc[0]
assert overall_10["status"] == "劣化警戒"
assert overall_10["return_delta"] < 0

thresholds = module.build_adaptive_threshold_recommendations(governance)
assert set(thresholds["mode"]) == {"shadow_only"}
assert (thresholds["recommended_value"] > thresholds["current_value"]).all()

# Healthy run sample.
all_ranked_rows = []
for index in range(120):
    all_ranked_rows.append({
        "code": f"{index:04d}",
        "sector33": f"業種{index % 30:02d}",
        "price_date": "2026-07-11",
        "score": 50 + index % 40,
        "rank": index + 1,
    })
all_ranked = pd.DataFrame(all_ranked_rows)
top100 = all_ranked.head(100).copy()
sector_momentum = pd.DataFrame({"sector33": [f"業種{i:02d}" for i in range(30)]})
run_health = module.build_run_health(
    "2026-07-11", all_ranked, top100, sector_momentum, leader_rows,
    [], 120, 120,
)
assert module.run_health_overall(run_health) == "PASS"
assert not (run_health["status"] == "FAIL").any()

failed_health = module.build_run_health(
    "2026-07-11", all_ranked.assign(sector33=""), top100, pd.DataFrame(), pd.DataFrame(),
    [{"error": "x"}] * 50, 120, 60,
)
assert module.run_health_overall(failed_health) == "FAIL"

plain = "\n".join(module.plain_governance_section(performance, governance, thresholds, run_health))
html = module.html_governance_section(performance, governance, thresholds, run_health)
assert "実績検証・運用品質" in plain
assert "劣化警戒" in plain
assert "shadow only" in plain
assert "実績検証・運用品質" in html
assert "PASS" in html

# Existing Action Priority remains intact.
action_row = pd.Series({
    "expectancy_score": 85,
    "expectancy_evidence_count": 12,
    "expectancy_confidence": "高",
    "score": 92,
    "rank": 3,
    "trading_value": 6_000_000_000,
    "volume_ratio": 3.5,
    "ma20_deviation": 0.10,
    "priority_labels": ["加速", "大型資金"],
    "priority_lifecycle_status": "定着",
    "priority_streak_days": 6,
})
action_result = module.action_priority_values(action_row, {"label": "強気"})
assert action_result["action_priority"] == "A"

# Workbook regression and required sheets.
with TemporaryDirectory() as tmpdir:
    report_path = str(Path(tmpdir) / "daily_report.xlsx")
    empty = pd.DataFrame()
    module.excel_report(
        report_path,
        {"実行日": "2026-07-11", "アプリ版": module.APP_VERSION},
        top100,
        sector_momentum,
        empty,
        leader_rows,
        snapshot,
        outcomes,
        performance,
        governance,
        thresholds,
        run_health,
        empty,
        empty,
        empty,
        empty,
        empty,
        empty,
        empty,
        empty,
        empty,
        empty,
        empty,
        [],
        empty,
    )
    workbook = load_workbook(report_path, read_only=True)
    required_sheets = {
        "Sector Leader History",
        "Sector Leader Outcomes",
        "Sector Leader Performance",
        "Signal Governance",
        "Adaptive Thresholds",
        "Run Health",
        "Action Priority",
    }
    assert required_sheets.issubset(set(workbook.sheetnames))
    workbook.close()

print("performance governance validation passed")
