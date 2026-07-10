from pathlib import Path
from tempfile import TemporaryDirectory
import importlib.util
import inspect
import sys

import pandas as pd
from openpyxl import load_workbook

spec = importlib.util.spec_from_file_location("momentum_main", "main.py")
module = importlib.util.module_from_spec(spec)
sys.modules[spec.name] = module
assert spec.loader is not None
spec.loader.exec_module(module)

assert module.APP_VERSION == "2026-07-11-dashboard-paper-portfolio-v15"
assert module.paper_target_exposure("強気", "PASS") == 0.80
assert module.paper_target_exposure("強気", "WARN") == 0.40
assert module.paper_target_exposure("強気", "FAIL") == 0.0
assert module.paper_target_exposure("過熱警戒", "PASS") == 0.30

health_pass = pd.DataFrame([
    {"check_name": "overall", "status": "PASS", "actual": "PASS", "expected": "PASS", "detail": "ok"},
])
health_fail = pd.DataFrame([
    {"check_name": "overall", "status": "FAIL", "actual": "FAIL", "expected": "PASS", "detail": "bad"},
])

leader_rows = []
for index in range(12):
    leader_rows.append({
        "overall_leader_rank": index + 1,
        "sector_leader_rank": index % 3 + 1,
        "sector33": f"業種{index % 4}",
        "sector_rank": index % 4 + 1,
        "sector_momentum_score": 80 - index,
        "sector_strength": "強い",
        "sector_rotation": "加速" if index < 6 else "主導",
        "sector_score_delta": 5.0,
        "code": f"{index + 1:04d}",
        "name": f"Leader{index + 1}",
        "close": 1000.0 + index * 50,
        "price_date": "2026-07-10",
        "momentum_rank": index + 1,
        "momentum_score": 90 - index,
        "sector_leader_score": 94 - index,
        "sector_leader_grade": "S" if index < 4 else "A",
        "sector_research_priority": "最優先" if index < 6 else "優先",
        "action_priority": "A" if index < 6 else "B",
        "action_score": 90 - index,
        "expectancy_score": 75.0,
        "expectancy_confidence": "高",
        "return_20d": 0.15,
        "return_60d": 0.25,
        "volume_ratio": 2.5,
        "trading_value": 3_000_000_000,
        "ma20_deviation": 0.10,
        "leader_reasons": "業種加速",
        "leader_cautions": "",
    })
leaders = pd.DataFrame(leader_rows)
empty_portfolio = pd.DataFrame(columns=module.PAPER_POSITION_COLUMNS)
empty_history = pd.DataFrame(columns=module.PAPER_TRADE_HISTORY_COLUMNS)
regime = {"label": "強気"}

plan = module.build_paper_trade_plan(
    "2026-07-11", leaders, empty_portfolio, empty_history, regime, health_pass,
)
assert not plan.empty
assert len(plan) <= module.PAPER_MAX_POSITIONS
assert (plan["quantity"] % module.PAPER_LOT_SIZE == 0).all()
assert (plan["quantity"] >= module.PAPER_LOT_SIZE).all()
assert plan["planned_value"].sum() <= module.PAPER_INITIAL_CAPITAL * 0.80 + 1
assert (plan["portfolio_weight"] <= module.PAPER_MAX_POSITION_WEIGHT + 1e-9).all()
sector_weights = plan.groupby("sector33")["planned_value"].sum() / module.PAPER_INITIAL_CAPITAL
assert (sector_weights <= module.PAPER_MAX_SECTOR_WEIGHT + 1e-9).all()
assert (plan["planned_risk"] <= module.PAPER_INITIAL_CAPITAL * module.PAPER_RISK_PER_TRADE + 1).all()

blocked_plan = module.build_paper_trade_plan(
    "2026-07-11", leaders, empty_portfolio, empty_history, regime, health_fail,
)
assert blocked_plan.empty

portfolio = module.apply_paper_trade_plan("2026-07-11", empty_portfolio, plan.head(2))
assert len(portfolio) == 2
assert set(portfolio["status"]) == {"OPEN"}
assert portfolio["position_id"].is_unique

# Stop-loss and take-profit exits are deterministic and close based.
exit_portfolio = portfolio.copy()
first_code = exit_portfolio.iloc[0]["code"]
second_code = exit_portfolio.iloc[1]["code"]
exit_prices = pd.DataFrame([
    {"code": first_code, "close": float(exit_portfolio.iloc[0]["stop_price"]) - 1, "price_date": "2026-07-14"},
    {"code": second_code, "close": float(exit_portfolio.iloc[1]["target_price"]) + 1, "price_date": "2026-07-14"},
])
marked, closed = module.mark_paper_positions("2026-07-14", exit_portfolio, exit_prices, {first_code, second_code})
assert marked.empty
assert set(closed["exit_reason"]) == {"STOP_LOSS", "TAKE_PROFIT"}
assert len(closed) == 2

trade_history = module.append_paper_trade_history(empty_history, closed)
assert len(trade_history) == 2
assert trade_history["position_id"].is_unique

totals = module.paper_portfolio_totals(empty_portfolio, trade_history)
assert totals["equity"] == module.PAPER_INITIAL_CAPITAL + trade_history["realized_pnl"].sum()

risk_budget = module.build_risk_budget(portfolio, module.paper_portfolio_totals(portfolio, empty_history), regime, health_pass)
assert not risk_budget.empty
assert set(risk_budget["status"]).issubset({"PASS", "FAIL"})
assert not (risk_budget["budget_type"] == "portfolio").empty

plain = "\n".join(module.plain_paper_portfolio_section(portfolio, plan, pd.DataFrame([module.paper_portfolio_totals(portfolio, empty_history) | {"drawdown": 0.0}]), risk_budget))
html = module.html_paper_portfolio_section(portfolio, plan, pd.DataFrame([module.paper_portfolio_totals(portfolio, empty_history) | {"drawdown": 0.0}]), risk_budget)
assert "ペーパーポートフォリオ" in plain
assert "実注文は行いません" in plain
assert "ペーパーポートフォリオ" in html
assert "実注文は行わない" in html

# Full state persistence, rerun idempotency, and equity history.
with TemporaryDirectory() as tmpdir:
    portfolio_path = str(Path(tmpdir) / "portfolio.csv")
    history_path = str(Path(tmpdir) / "trades.csv")
    equity_path = str(Path(tmpdir) / "equity.csv")
    ranked = leaders.rename(columns={"momentum_rank": "rank", "momentum_score": "score"})[["code", "name", "sector33", "close", "price_date", "rank", "score"]].copy()
    result1 = module.run_paper_portfolio(
        "2026-07-11", ranked, leaders, regime, health_pass,
        portfolio_path, history_path, equity_path,
    )
    assert Path(portfolio_path).exists()
    assert Path(history_path).exists()
    assert Path(equity_path).exists()
    assert len(result1["portfolio"]) <= module.PAPER_MAX_POSITIONS
    result2 = module.run_paper_portfolio(
        "2026-07-11", ranked, leaders, regime, health_pass,
        portfolio_path, history_path, equity_path,
    )
    assert result2["portfolio"]["position_id"].is_unique
    assert len(result2["equity_history"]) == 1

# Existing governance and Action Priority behavior remain intact.
degradation_rows = []
for index in range(40):
    recent = index >= 20
    degradation_rows.append({
        "signal_date": (pd.Timestamp("2025-01-01") + pd.Timedelta(days=index)).date().isoformat(),
        "entry_price_date": "2025-01-01",
        "exit_price_date": "2025-02-01",
        "code": f"{index:04d}",
        "name": f"Stock{index}",
        "sector33": "電気機器",
        "sector_research_priority": "最優先",
        "sector_leader_grade": "S",
        "sector_rotation": "加速",
        "sector_leader_score": 90.0,
        "horizon_days": 10,
        "entry_close": 100.0,
        "exit_close": 96.0 if recent else 106.0,
        "forward_return": -0.04 if recent else 0.06,
        "win": not recent,
        "calendar_days": 14,
    })
governance = module.build_signal_governance(pd.DataFrame(degradation_rows))
assert governance[(governance["scope_type"] == "overall") & (governance["horizon_days"] == 10)].iloc[0]["status"] == "劣化警戒"

action_row = pd.Series({
    "expectancy_score": 85,
    "expectancy_evidence_count": 12,
    "expectancy_confidence": "高",
    "score": 92,
    "rank": 3,
    "trading_value": 6_000_000_000,
    "volume_ratio": 3.5,
    "ma20_deviation": 0.10,
    "priority_labels": ["加速", "大型資金"],
    "priority_lifecycle_status": "定着",
    "priority_streak_days": 6,
})
assert module.action_priority_values(action_row, {"label": "強気"})["action_priority"] == "A"

# Workbook regression with new paper sheets.
with TemporaryDirectory() as tmpdir:
    report_path = str(Path(tmpdir) / "daily_report.xlsx")
    empty = pd.DataFrame()
    signature = inspect.signature(module.excel_report)
    values = {name: empty for name in signature.parameters}
    values.update({
        "path": report_path,
        "summary": {"実行日": "2026-07-11", "アプリ版": module.APP_VERSION},
        "top100": leaders.head(5),
        "sector_leaders": leaders,
        "run_health": health_pass,
        "paper_portfolio": portfolio,
        "paper_trade_plan": plan,
        "paper_trade_history": trade_history,
        "paper_risk_budget": risk_budget,
        "paper_performance": pd.DataFrame([module.paper_portfolio_totals(portfolio, empty_history) | {"drawdown": 0.0}]),
        "errors": [],
    })
    module.excel_report(**values)
    workbook = load_workbook(report_path, read_only=True)
    required = {
        "Paper Portfolio", "Paper Trade Plan", "Paper Trade History",
        "Risk Budget", "Paper Performance", "Run Health", "Action Priority",
    }
    assert required.issubset(set(workbook.sheetnames))
    workbook.close()

print("paper portfolio and risk validation passed")
