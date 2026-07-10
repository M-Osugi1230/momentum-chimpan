from pathlib import Path
from tempfile import TemporaryDirectory
import importlib.util
import inspect
import sys

import pandas as pd
from openpyxl import load_workbook

spec = importlib.util.spec_from_file_location("momentum_main", "main.py")
module = importlib.util.module_from_spec(spec)
sys.modules[spec.name] = module
assert spec.loader is not None
spec.loader.exec_module(module)

assert module.APP_VERSION == "2026-07-11-dashboard-release-readiness-v16"
assert module.EXECUTION_MODE == "RESEARCH_AND_PAPER_ONLY"

health_pass = pd.DataFrame([
    {"check_name": "overall", "status": "PASS", "actual": "PASS", "expected": "PASS", "detail": "ok"},
])
health_fail = pd.DataFrame([
    {"check_name": "overall", "status": "FAIL", "actual": "FAIL", "expected": "PASS", "detail": "bad"},
])
performance = pd.DataFrame([
    {"group_type": "overall", "group_value": "ALL", "horizon_days": 10, "count": 35, "win_rate": 0.60, "average_return": 0.04, "median_return": 0.03, "best_return": 0.20, "worst_return": -0.08, "average_leader_score": 82.0},
])
governance_stable = pd.DataFrame([
    {"scope_type": "overall", "scope_value": "ALL", "horizon_days": 10, "evidence_count": 35, "recent_count": 20, "baseline_average_return": 0.03, "recent_average_return": 0.04, "return_delta": 0.01, "baseline_win_rate": 0.55, "recent_win_rate": 0.60, "win_rate_delta": 0.05, "status": "安定", "health_score": 75, "recommendation": "維持"},
])
paper_performance_good = pd.DataFrame([{
    "date": "2026-07-11",
    "initial_capital": module.PAPER_INITIAL_CAPITAL,
    "cash": 5_000_000.0,
    "invested_cost": 5_000_000.0,
    "market_value": 5_500_000.0,
    "equity": 10_500_000.0,
    "realized_pnl": 200_000.0,
    "unrealized_pnl": 300_000.0,
    "exposure_ratio": 5_500_000 / 10_500_000,
    "peak_equity": 10_700_000.0,
    "drawdown": -0.05,
    "open_positions": 5,
    "closed_trades": 24,
    "win_rate": 0.625,
}])
trade_history = pd.DataFrame([{"position_id": f"p{i}", "realized_pnl": 10_000 if i < 15 else -5_000} for i in range(24)])
risk_pass = pd.DataFrame([
    {"budget_type": "portfolio", "label": "投資比率", "current_value": 0.52, "limit_value": 0.80, "utilization": 0.65, "status": "PASS", "detail": "ok"},
])

ready = module.build_release_readiness(
    health_pass, governance_stable, performance, paper_performance_good, trade_history, risk_pass,
)
assert module.release_status_value(ready) == "READY_FOR_MANUAL_REVIEW"
assert ready["passed"].all()
assert set(ready["execution_mode"]) == {module.EXECUTION_MODE}
alerts_ready = module.build_operational_alerts(ready)
assert set(alerts_ready["severity"]) == {"INFO"}

paper_no_trades = paper_performance_good.copy()
paper_no_trades.loc[:, "equity"] = module.PAPER_INITIAL_CAPITAL
paper_no_trades.loc[:, "win_rate"] = None
research = module.build_release_readiness(
    health_pass, governance_stable, performance, paper_no_trades, pd.DataFrame(), risk_pass,
)
assert module.release_status_value(research) == "RESEARCH"

hold_health = module.build_release_readiness(
    health_fail, governance_stable, performance, paper_performance_good, trade_history, risk_pass,
)
assert module.release_status_value(hold_health) == "HOLD"
health_alerts = module.build_operational_alerts(hold_health)
assert "P0" in set(health_alerts["severity"])
assert "Run Health" in set(health_alerts["title"])

bad_drawdown = paper_performance_good.copy()
bad_drawdown.loc[:, "drawdown"] = -0.12
hold_drawdown = module.build_release_readiness(
    health_pass, governance_stable, performance, bad_drawdown, trade_history, risk_pass,
)
assert module.release_status_value(hold_drawdown) == "HOLD"
assert "最大ドローダウン" in set(module.build_operational_alerts(hold_drawdown)["title"])

governance_bad = governance_stable.copy()
governance_bad.loc[:, "status"] = "劣化警戒"
hold_degradation = module.build_release_readiness(
    health_pass, governance_bad, performance, paper_performance_good, trade_history, risk_pass,
)
assert module.release_status_value(hold_degradation) == "HOLD"
assert "P1" in set(module.build_operational_alerts(hold_degradation)["severity"])

with TemporaryDirectory() as tmpdir:
    root = Path(tmpdir)
    state_paths = {}
    for index, name in enumerate(["ranking", "temperature", "paper"]):
        state_path = root / f"{name}.csv"
        pd.DataFrame([{"code": f"{index:04d}", "value": index + 1}]).to_csv(state_path, index=False)
        state_paths[name] = str(state_path)
    state_paths["missing"] = str(root / "missing.csv")

    inventory = module.build_state_inventory(state_paths)
    assert len(inventory) == 4
    assert int((inventory["status"] == "OK").sum()) == 3
    assert int((inventory["status"] == "MISSING").sum()) == 1
    assert inventory[inventory["status"] == "OK"]["sha256"].str.len().eq(64).all()

    snapshots = module.snapshot_state_files("2026-07-11", state_paths, str(root / "snapshots"))
    assert int((snapshots["status"] == "SNAPSHOT_CREATED").sum()) == 3
    for snapshot_path in snapshots[snapshots["status"] == "SNAPSHOT_CREATED"]["snapshot_path"]:
        assert Path(snapshot_path).exists()

    alerts = module.build_operational_alerts(ready)
    audit = module.build_execution_audit("2026-07-11", ready, alerts, inventory, snapshots, health_pass)
    assert audit.iloc[0]["execution_mode"] == module.EXECUTION_MODE
    assert len(audit.iloc[0]["manifest_sha256"]) == 64
    audit_path = str(root / "audit.csv")
    first = module.append_execution_audit(audit_path, audit)
    second = module.append_execution_audit(audit_path, audit)
    assert len(first) == 1
    assert len(second) == 1

    plain = "\n".join(module.plain_release_readiness_section(ready, alerts, inventory))
    html = module.html_release_readiness_section(ready, alerts, inventory)
    assert "READY_FOR_MANUAL_REVIEW" in plain
    assert "自動発注は無効" in plain
    assert "READY_FOR_MANUAL_REVIEW" in html
    assert "自動発注は無効" in html

    report_path = str(root / "daily_report.xlsx")
    empty = pd.DataFrame()
    signature = inspect.signature(module.excel_report)
    values = {name: empty for name in signature.parameters}
    values.update({
        "path": report_path,
        "summary": {"実行日": "2026-07-11", "アプリ版": module.APP_VERSION},
        "run_health": health_pass,
        "paper_trade_history": trade_history,
        "paper_risk_budget": risk_pass,
        "paper_performance": paper_performance_good,
        "release_readiness": ready,
        "operational_alerts": alerts,
        "state_inventory": inventory,
        "state_snapshots": snapshots,
        "execution_audit": audit,
        "errors": [],
    })
    module.excel_report(**values)
    workbook = load_workbook(report_path, read_only=True)
    required = {
        "Release Readiness", "Operational Alerts", "State Inventory",
        "State Snapshots", "Execution Audit", "Paper Performance", "Run Health",
    }
    assert required.issubset(set(workbook.sheetnames))
    workbook.close()

# Existing paper risk and action-priority behavior remain unchanged.
assert module.paper_target_exposure("強気", "PASS") == 0.80
assert module.paper_target_exposure("弱気", "PASS") == 0.20
assert module.paper_target_exposure("強気", "FAIL") == 0.0

action_row = pd.Series({
    "expectancy_score": 85,
    "expectancy_evidence_count": 12,
    "expectancy_confidence": "高",
    "score": 92,
    "rank": 3,
    "trading_value": 6_000_000_000,
    "volume_ratio": 3.5,
    "ma20_deviation": 0.10,
    "priority_labels": ["加速", "大型資金"],
    "priority_lifecycle_status": "定着",
    "priority_streak_days": 6,
})
assert module.action_priority_values(action_row, {"label": "強気"})["action_priority"] == "A"

print("release readiness and operations audit validation passed")
