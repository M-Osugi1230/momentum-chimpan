from __future__ import annotations

import sys
from pathlib import Path
from tempfile import TemporaryDirectory

import pandas as pd
from openpyxl import Workbook, load_workbook

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

import daily_research_focus as focus


policy = focus.load_policy(ROOT / focus.POLICY_PATH)
focus.validate_policy(policy)
assert policy["limits"]["maximum_A_candidates"] == 5
assert policy["limits"]["maximum_daily_action_list"] == 10
assert policy["governance"]["preserve_paper_execution"] is True
assert policy["governance"]["automatic_strategy_change"] is False
assert policy["governance"]["production_state_mutations"] == []

priority_rows = []
top_rows = []
for position in range(1, 15):
    code = str(1000 + position)
    if position <= 7:
        action_priority = "A"
        action_score = 100 - position
        lifecycle = "継続"
    elif position <= 11:
        action_priority = "B"
        action_score = 82 - position
        lifecycle = "定着"
    elif position == 12:
        action_priority = "C"
        action_score = 60
        lifecycle = "継続"
    elif position == 13:
        action_priority = "C"
        action_score = 52
        lifecycle = "初登場"
    else:
        action_priority = "見送り"
        action_score = 47
        lifecycle = "初登場"
    priority_rows.append({
        "code": code,
        "name": f"Candidate {position}",
        "momentum_rank": position,
        "momentum_score": 95 - position,
        "priority_labels": "初動 / 加速" if position <= 3 else "継続",
        "lifecycle_status": lifecycle,
        "expectancy_score": 90 - position,
        "expectancy_confidence": "中",
        "expectancy_evidence_count": 10,
        "market_regime": "やや強気",
        "action_priority": action_priority,
        "action_score": action_score,
        "positive_reasons": "既存の強さ / 流動性良好",
        "caution_reasons": "20日線乖離を確認",
        "liquidity_check": "流動性良好（売買代金10億円以上）",
        "overheat_check": "過熱注意なし",
        "return_20d": 0.20,
        "ma20_deviation": 0.10,
        "volume_ratio": 2.5,
        "trading_value": 2_000_000_000,
        "data_quality_grade": "A" if position != 13 else "C",
        "data_quality_warnings": "" if position != 13 else "分析項目を追加確認",
        "data_quality_reason_codes": "" if position != 13 else "MISSING_ANALYTICAL_FIELD",
        "data_quality_eligible_for_a": position != 13,
    })
    top_rows.append({
        "code": code,
        "rank": position,
        "score": 95 - position,
        "sector33": "電気機器",
        "is_new_entry": position in {1, 8, 13},
        "rank_change": 30 if position == 1 else 5 if position == 2 else 0,
        "is_rising_fast": position == 1,
        "is_best_rank": position in {1, 2},
        "top30_streak": 5 if position <= 5 else 1,
        "relative_strength_grade": "A" if position <= 5 else "B",
        "relative_strength_rank": position,
        "market_relative_20d": 0.05,
        "sector_relative_20d": 0.03,
        "return_5d": 0.08,
        "return_20d": 0.20,
        "volume_ratio": 2.5,
        "trading_value": 2_000_000_000,
        "ma20_deviation": 0.10,
        "data_quality_grade": "A" if position != 13 else "C",
        "data_quality_warnings": "" if position != 13 else "分析項目を追加確認",
        "data_quality_reason_codes": "" if position != 13 else "MISSING_ANALYTICAL_FIELD",
        "data_quality_eligible_for_a": position != 13,
    })

action = pd.DataFrame(priority_rows)
top100 = pd.DataFrame(top_rows)
original_ranks = action.set_index("code")["momentum_rank"].copy()
original_scores = action.set_index("code")["momentum_score"].copy()
result = focus.attach_daily_focus(
    action,
    top100,
    policy_path=ROOT / focus.POLICY_PATH,
)

assert int((result["research_bucket"] == "A").sum()) == 5
assert int(result["daily_action_list"].sum()) == 10
assert result[result["daily_action_list"]]["daily_action_rank"].tolist() == list(range(1, 11))
assert result["explanation_complete"].all()
assert not result["why_today"].astype(str).str.strip().eq("").any()
assert not result["what_changed"].astype(str).str.strip().eq("").any()
assert not result["risk_summary"].astype(str).str.strip().eq("").any()
assert not result["next_research_questions"].astype(str).str.strip().eq("").any()
assert set(result[result["research_bucket"] == "A"]["code"]) == {"1001", "1002", "1003", "1004", "1005"}
for code in ("1006", "1007"):
    row = result.set_index("code").loc[code]
    assert row["research_bucket"] == "B"
    assert row["action_priority"] == "B"
    assert "A候補上限5件" in row["focus_adjustment_reason"]
assert result.set_index("code").loc["1012", "research_bucket"] == "C"
assert result.set_index("code").loc["1013", "research_bucket"] == "Watch"
assert result.set_index("code").loc["1014", "research_bucket"] == "Watch"
assert "Top100新規ランクイン" in result.set_index("code").loc["1001", "why_today"]
assert "前回比+30位の急上昇" in result.set_index("code").loc["1001", "what_changed"]
assert "Data Quality C" in result.set_index("code").loc["1013", "risk_summary"]
assert "最新決算" in result.set_index("code").loc["1001", "next_research_questions"]
assert "出来高急増" in result.set_index("code").loc["1001", "next_research_questions"]

reindexed = result.set_index("code")
pd.testing.assert_series_equal(
    reindexed.loc[original_ranks.index, "momentum_rank"],
    original_ranks,
    check_names=False,
)
pd.testing.assert_series_equal(
    reindexed.loc[original_scores.index, "momentum_score"],
    original_scores,
    check_names=False,
)

fields = focus.summary_fields(result)
assert fields["Daily Focus A"] == 5
assert fields["Daily Focus B"] == 6
assert fields["Daily Focus C"] == 1
assert fields["Daily Focus Watch"] == 2
assert fields["Daily Focus Skip"] == 0
assert fields["Daily Action List"] == 10
assert fields["Daily Focus説明不足"] == 0
assert fields["Daily Focus A上限超過"] == 0

selected = focus.action_list(result)
assert len(selected) == 10
assert set(selected["research_bucket"]).issubset({"A", "B"})

plain = "\n".join(focus.plain_section(result))
assert "【今日の結論・Daily Action List】" in plain
assert "A 5件" in plain
assert "詳細調査対象 10件" in plain
assert "今日の理由：" in plain
assert "次の確認：" in plain
html = focus.html_section(result)
assert "今日の結論・Daily Action List" in html
assert "A 5" in html
assert "詳細調査 10件" in html
assert "次の確認：" in html

with TemporaryDirectory() as temporary:
    workbook_path = Path(temporary) / "daily.xlsx"
    workbook = Workbook()
    workbook.active.title = "Summary"
    workbook.save(workbook_path)
    focus.patch_workbook(workbook_path, result)
    checked = load_workbook(workbook_path, data_only=True)
    assert "Daily Action List" in checked.sheetnames
    sheet = checked["Daily Action List"]
    values = [cell.value for row in sheet.iter_rows() for cell in row]
    assert "Detailed research list" in values
    assert "A cap violations" in values
    assert "Paper execution mutation" in values
    assert "NONE" in values
    assert "why_today" in values
    assert "next_research_questions" in values
    assert "Candidate 1" in values

source = (ROOT / "daily_runner.py").read_text(encoding="utf-8")
assert "daily_research_focus.attach_daily_focus" in source
assert "daily_research_focus.patch_workbook" in source
assert "daily_research_focus.plain_section" in source
assert "daily_research_focus.html_section" in source
assert "paper_execution_mutation=disabled" in source

print("daily research focus validation passed")
