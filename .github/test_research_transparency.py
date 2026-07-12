from __future__ import annotations

import sys
from pathlib import Path
from tempfile import TemporaryDirectory

from openpyxl import Workbook, load_workbook

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

import daily_runner
import research_evidence_catalog as evidence_catalog
import research_transparency as transparency


snapshot = transparency.load_snapshot(
    ROOT / "research" / "evidence_catalog.yaml",
    ROOT,
)
assert snapshot["catalog_health"] == "PASS"
assert snapshot["production_weight_points"] == 15
assert snapshot["decision"] == evidence_catalog.HOLD_DECISION
assert snapshot["historical_consensus"] == "CONFLICTED_TIME_UNSTABLE"
assert snapshot["research_status"] == "UNRESOLVED"
assert snapshot["governing_study_status"] == "ACCUMULATING"
assert snapshot["automatic_weight_change_allowed"] is False
assert snapshot["automatic_strategy_change_allowed"] is False
assert len(snapshot["studies"]) == 4

summary = daily_runner.enrich_summary({"実行日": "2026-07-13"}, snapshot)
assert summary["実行日"] == "2026-07-13"
assert summary["出来高倍率配点"] == 15
assert summary["研究判断"] == evidence_catalog.HOLD_DECISION
assert summary["歴史エビデンス"] == "CONFLICTED_TIME_UNSTABLE"
assert summary["Forward Evidence"] == "ACCUMULATING"
assert summary["自動配点変更"] == "DISABLED"

plain = "\n".join(transparency.plain_section(snapshot))
assert "研究エビデンスの現在地" in plain
assert "現行配点 15点（据え置き）" in plain
assert "CONFLICTED_TIME_UNSTABLE" in plain
assert "Forward Evidence: ACCUMULATING" in plain
assert "自動配点変更・自動戦略変更は無効" in plain

plain_body = "HEADER\n【Market Temperature】\nDETAIL"
plain_patched = daily_runner.insert_plain_section(
    plain_body,
    transparency.plain_section(snapshot),
)
assert plain_patched.index("研究エビデンスの現在地") < plain_patched.index(
    "【Market Temperature】"
)
assert plain_patched.count("【Market Temperature】") == 1

html = transparency.html_section(snapshot)
assert "RESEARCH EVIDENCE" in html
assert "15点のまま維持" in html
assert "CONFLICTED_TIME_UNSTABLE" in html
assert "ACCUMULATING" in html
assert "自動配点変更なし" in html

html_body = f"<html><body>HEADER{daily_runner.HTML_MARKER}DETAIL</body></html>"
html_patched = daily_runner.insert_html_section(html_body, html)
assert html_patched.index("RESEARCH EVIDENCE") < html_patched.index(
    "Market Temperature"
)
assert html_patched.count("RESEARCH EVIDENCE") == 1

fallback = transparency.load_snapshot(
    ROOT / "research" / "definitely-missing-catalog.yaml",
    ROOT,
)
assert fallback["catalog_health"] == "WARN"
assert fallback["production_weight_points"] == 15
assert fallback["decision"] == evidence_catalog.HOLD_DECISION
assert fallback["automatic_weight_change_allowed"] is False
assert "SAFE HOLD" in transparency.html_section(fallback)

with TemporaryDirectory() as temporary:
    workbook_path = Path(temporary) / "report.xlsx"
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Summary"
    summary_sheet.append(["実行日", "出来高倍率配点"])
    summary_sheet.append(["2026-07-13", 15])
    workbook.save(workbook_path)

    transparency.patch_workbook(workbook_path, snapshot)
    reopened = load_workbook(workbook_path, data_only=True)
    assert reopened.sheetnames[0] == "Summary"
    assert reopened.sheetnames[1] == transparency.SHEET_NAME
    evidence_sheet = reopened[transparency.SHEET_NAME]
    assert evidence_sheet.freeze_panes == "A2"
    values = [cell.value for row in evidence_sheet.iter_rows() for cell in row]
    assert "Current Decision" in values
    assert "volume-component-forward-evidence-v1" in values
    assert "NOT_SUPPORTED" in values
    assert "ACCUMULATING" in values

workflow_text = (
    ROOT / ".github" / "workflows" / "daily.yml"
).read_text(encoding="utf-8")
assert "python daily_runner.py" in workflow_text
assert "python main.py 2>&1 | tee output/run.log" not in workflow_text
assert "Snapshot governed strategy fingerprint before report" in workflow_text
assert workflow_text.index("Snapshot governed strategy fingerprint before report") < workflow_text.index(
    "python daily_runner.py"
)

print("research transparency dashboard validation passed")
