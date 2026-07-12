from __future__ import annotations

import copy
import json
import shutil
import sys
from pathlib import Path
from tempfile import TemporaryDirectory

import pandas as pd
from openpyxl import Workbook

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

import email_delivery
import live_session_readiness as readiness


FINGERPRINT = "a" * 64
REPORT_DATE = "2026-07-13"


def write_json(path: Path, payload: dict) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def build_artifact(root: Path, email_status: str = "SMTP_ACCEPTED") -> Path:
    artifact = root / "momentum-operations-123"
    output = artifact / "output"
    data = artifact / "data"
    output.mkdir(parents=True, exist_ok=True)
    data.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    summary = workbook.active
    summary.title = "Summary"
    summary.append(["実行日", "状態更新実行", "Data Quality評価件数"])
    summary.append([REPORT_DATE, "YES", 2])

    top100 = workbook.create_sheet("Momentum Top100")
    top_columns = [
        "rank", "code", "name", "score", "data_quality_grade",
        "data_quality_score", "data_quality_eligible_for_a",
        "data_quality_reason_codes", "data_quality_warnings",
    ]
    top100.append(top_columns)
    top100.append([1, "1111", "Candidate A", 90, "A", 100, True, "", ""])
    top100.append([2, "2222", "Candidate B", 85, "B", 95, True, "BLANK_SECTOR", "業種空欄"])

    action = workbook.create_sheet("Action Priority")
    action_columns = [
        "code", "name", "sector33", "research_bucket", "daily_action_list",
        "daily_action_rank", "action_priority", "action_priority_before_quality",
        "action_priority_before_daily_focus", "momentum_rank", "momentum_score",
        "action_score", "expectancy_score", "expectancy_confidence",
        "lifecycle_status", "market_regime", "relative_strength_grade",
        "data_quality_grade", "data_quality_reason_codes", "why_today",
        "what_changed", "risk_summary", "next_research_questions",
        "focus_adjustment_reason", "daily_focus_version",
    ]
    action.append(action_columns)
    rows = [
        {
            "code": "1111", "name": "Candidate A", "sector33": "電気機器",
            "research_bucket": "A", "daily_action_list": True, "daily_action_rank": 1,
            "action_priority": "A", "action_priority_before_quality": "A",
            "action_priority_before_daily_focus": "A", "momentum_rank": 1,
            "momentum_score": 90, "action_score": 92, "expectancy_score": 75,
            "expectancy_confidence": "中", "lifecycle_status": "継続",
            "market_regime": "やや強気", "relative_strength_grade": "A",
            "data_quality_grade": "A", "data_quality_reason_codes": "",
            "why_today": "Top100新規と相対強度", "what_changed": "前回比+20位",
            "risk_summary": "Data Quality A / 過熱注意なし",
            "next_research_questions": "最新決算と材料継続性を確認",
            "focus_adjustment_reason": "変更なし",
            "daily_focus_version": "2026-07-12-daily-research-focus-v1",
        },
        {
            "code": "2222", "name": "Candidate B", "sector33": "機械",
            "research_bucket": "B", "daily_action_list": True, "daily_action_rank": 2,
            "action_priority": "B", "action_priority_before_quality": "B",
            "action_priority_before_daily_focus": "B", "momentum_rank": 2,
            "momentum_score": 85, "action_score": 75, "expectancy_score": 65,
            "expectancy_confidence": "蓄積中", "lifecycle_status": "初登場",
            "market_regime": "やや強気", "relative_strength_grade": "B",
            "data_quality_grade": "B", "data_quality_reason_codes": "BLANK_SECTOR",
            "why_today": "出来高増加", "what_changed": "新規ランクイン",
            "risk_summary": "Data Quality B / 業種空欄",
            "next_research_questions": "適時開示とチャートを確認",
            "focus_adjustment_reason": "変更なし",
            "daily_focus_version": "2026-07-12-daily-research-focus-v1",
        },
    ]
    for row in rows:
        action.append([row.get(column) for column in action_columns])

    workbook.create_sheet("Daily Action List")
    workbook.create_sheet("Data Quality")
    workbook.create_sheet("Research Evidence")
    workbook.save(output / "daily_report.xlsx")

    write_json(data / "operations_heartbeat.json", {
        "workflow_status": "SUCCESS",
        "state_update_executed": True,
        "report_date": REPORT_DATE,
        "current_day_price_ratio": 0.99,
        "market_data_freshness": "FRESH",
        "research_only": True,
    })
    write_json(data / "strategy_fingerprint.json", {
        "strategy_fingerprint": FINGERPRINT,
    })
    ranking = pd.DataFrame([
        {
            "date": REPORT_DATE,
            "rank": 1,
            "code": "1111",
            "name": "Candidate A",
            "score": 90,
            "strategy_fingerprint": FINGERPRINT,
        },
        {
            "date": REPORT_DATE,
            "rank": 2,
            "code": "2222",
            "name": "Candidate B",
            "score": 85,
            "strategy_fingerprint": FINGERPRINT,
        },
    ])
    ranking.to_csv(data / "momentum_daily_ranking.csv", index=False)
    write_json(output / "evidence_stamp_audit.json", {
        "report_date": REPORT_DATE,
        "strategy_fingerprint": FINGERPRINT,
        "stamped_rows": 2,
        "research_only": True,
    })
    write_json(output / "recovery_snapshot_audit.json", {
        "snapshot_date": REPORT_DATE,
        "status": "SEALED",
        "complete": True,
        "research_only": True,
    })
    write_json(output / "state_maintenance.json", {
        "validation_status": "PASS",
        "validation_failures": 0,
    })

    receipt = email_delivery.build_receipt(
        status=email_status,
        summary={"実行日": REPORT_DATE},
        sender="sender@example.com" if email_status != "SKIPPED_SECRETS_MISSING" else "",
        recipient_text="recipient@example.com" if email_status != "SKIPPED_SECRETS_MISSING" else "",
        started_at_utc="2026-07-13T08:00:00+00:00",
        completed_at_utc="2026-07-13T08:00:05+00:00",
        error_class="RuntimeError" if email_status == "FAILED" else "",
    )
    if email_status == "SMTP_ACCEPTED":
        # build_receipt reads the password from the environment. Set the flag and
        # rebuild the envelope to create a deterministic valid synthetic receipt.
        receipt["secret_configuration_complete"] = True
        receipt.pop("status_sha256", None)
        receipt.pop("receipt_fingerprint", None)
        substantive = dict(receipt)
        receipt["receipt_fingerprint"] = email_delivery.canonical_hash(substantive)
        receipt["status_sha256"] = email_delivery.canonical_hash(receipt)
    write_json(output / "email_delivery_receipt.json", receipt)
    return artifact


def run(artifact: Path, conclusion: str = "success", event: str = "schedule") -> dict:
    return readiness.build_readiness(
        artifact,
        source_run_id="123",
        source_run_url="https://example.test/runs/123",
        upstream_conclusion=conclusion,
        upstream_event=event,
        head_sha="b" * 40,
        created_at_utc="2026-07-13T07:45:00Z",
        updated_at_utc="2026-07-13T08:10:00Z",
        generated_at_utc="2026-07-13T08:11:00+00:00",
    )


with TemporaryDirectory() as temporary:
    root = Path(temporary)
    artifact = build_artifact(root)
    passed = run(artifact)
    assert readiness.validate_readiness(passed) == []
    assert passed["readiness_state"] == "PASS"
    assert passed["fail_count"] == 0
    assert passed["review_required_count"] == 0
    assert passed["eligible_for_forward_evidence"] is True
    assert passed["eligible_for_priority_outcome_ingestion"] is True
    assert passed["smtp_acceptance_only"] is True
    assert passed["inbox_delivery_claimed"] is False
    gates = {item["name"]: item for item in passed["gates"]}
    assert gates["ranking_history"]["metrics"]["report_date_row_count"] == 2
    assert gates["ranking_history"]["metrics"]["duplicate_row_count"] == 0
    assert gates["data_quality"]["metrics"]["quality_coverage"] == 1.0
    assert gates["daily_research_focus"]["metrics"]["priority_a_count"] == 1
    assert gates["daily_research_focus"]["metrics"]["daily_action_list_count"] == 2
    assert gates["smtp_acceptance"]["status"] == "PASS"
    assert gates["priority_outcome_ingestion"]["metrics"]["extractable_decision_count"] == 2
    markdown = readiness.markdown_report(passed)
    assert "State: **PASS**" in markdown
    assert "Forward Evidence eligible: **True**" in markdown
    assert "Inbox delivery claimed: **False**" in markdown

    skipped_root = root / "skipped"
    skipped = build_artifact(skipped_root, email_status="SKIPPED_SECRETS_MISSING")
    review = run(skipped)
    assert readiness.validate_readiness(review) == []
    assert review["readiness_state"] == "REVIEW_REQUIRED"
    review_gates = {item["name"]: item for item in review["gates"]}
    assert review_gates["smtp_acceptance"]["status"] == "REVIEW_REQUIRED"
    assert review["eligible_for_forward_evidence"] is True
    assert review["eligible_for_priority_outcome_ingestion"] is True

    low_coverage_root = root / "low-coverage"
    low_coverage = build_artifact(low_coverage_root)
    heartbeat_path = next(low_coverage.rglob("operations_heartbeat.json"))
    heartbeat = json.loads(heartbeat_path.read_text(encoding="utf-8"))
    heartbeat["current_day_price_ratio"] = 0.95
    write_json(heartbeat_path, heartbeat)
    coverage_review = run(low_coverage)
    assert coverage_review["readiness_state"] == "REVIEW_REQUIRED"
    coverage_gate = {item["name"]: item for item in coverage_review["gates"]}["market_data_coverage"]
    assert coverage_gate["status"] == "REVIEW_REQUIRED"

    failed_root = root / "failed"
    failed_artifact = build_artifact(failed_root)
    ranking_path = next(failed_artifact.rglob("momentum_daily_ranking.csv"))
    ranking_path.unlink()
    failed = run(failed_artifact)
    assert readiness.validate_readiness(failed) == []
    assert failed["readiness_state"] == "FAIL"
    failed_gates = {item["name"]: item for item in failed["gates"]}
    assert failed_gates["artifact_completeness"]["status"] == "FAIL"
    assert failed_gates["ranking_history"]["status"] == "FAIL"
    assert failed["eligible_for_forward_evidence"] is False

    upstream_failure_root = root / "upstream-failure"
    upstream_artifact = build_artifact(upstream_failure_root)
    upstream_failed = run(upstream_artifact, conclusion="failure")
    assert upstream_failed["readiness_state"] == "FAIL"
    assert {item["name"]: item for item in upstream_failed["gates"]}["upstream_workflow"]["status"] == "FAIL"

    tampered = copy.deepcopy(passed)
    tampered["automatic_strategy_change"] = True
    issues = readiness.validate_readiness(tampered)
    assert "automatic_strategy_change must be false" in issues
    assert "status_sha256 mismatch" in issues

    output_paths = readiness.write_outputs(passed, root / "output")
    assert Path(output_paths["json"]).is_file()
    assert Path(output_paths["markdown"]).is_file()
    written = json.loads(Path(output_paths["json"]).read_text(encoding="utf-8"))
    assert readiness.validate_readiness(written) == []

workflow_text = (ROOT / ".github/workflows/live-session-readiness.yml").read_text(encoding="utf-8")
assert "Daily Momentum Report" in workflow_text
assert "workflow_run" in workflow_text
assert "actions: read" in workflow_text
assert "contents: read" in workflow_text
assert "live_session_readiness.py build" in workflow_text
assert "actions/download-artifact@v4" in workflow_text
assert "actions/upload-artifact@v4" in workflow_text
assert ("git" + " push") not in workflow_text
assert ("contents:" + " write") not in workflow_text
assert ("EMAIL_" + "APP_PASSWORD") not in workflow_text

print("live session readiness validation passed")
