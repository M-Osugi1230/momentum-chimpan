from __future__ import annotations

import sys
from pathlib import Path
from tempfile import TemporaryDirectory

import pandas as pd
from openpyxl import Workbook, load_workbook

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

import data_quality
import main


policy = data_quality.load_policy(ROOT / data_quality.POLICY_PATH)
data_quality.validate_policy(policy)
assert policy["priority_boundary"]["preserve_momentum_score"] is True
assert policy["priority_boundary"]["preserve_momentum_rank"] is True
assert policy["governance"]["automatic_strategy_change"] is False
assert policy["governance"]["automatic_weight_change"] is False
assert policy["governance"]["production_state_mutations"] == []

rows = pd.DataFrame([
    {
        "date": "2026-07-13",
        "rank": 1,
        "code": "1111",
        "name": "Quality A",
        "sector33": "電気機器",
        "close": 1000.0,
        "prev_close": 980.0,
        "volume": 1_000_000,
        "trading_value": 1_000_000_000,
        "return_5d": 0.08,
        "return_20d": 0.20,
        "volume_ratio": 2.0,
        "ma20": 950.0,
        "ma60": 900.0,
        "ma20_deviation": 0.05,
        "price_date": "2026-07-13",
        "score": 90,
    },
    {
        "date": "2026-07-13",
        "rank": 2,
        "code": "2222",
        "name": "Quality B",
        "sector33": "",
        "close": 2000.0,
        "prev_close": 1980.0,
        "volume": 500_000,
        "trading_value": 1_000_000_000,
        "return_5d": 0.05,
        "return_20d": 0.15,
        "volume_ratio": 2.5,
        "ma20": 1900.0,
        "ma60": 1800.0,
        "ma20_deviation": 0.05,
        "price_date": "2026-07-13",
        "score": 85,
    },
    {
        "date": "2026-07-13",
        "rank": 3,
        "code": "3333",
        "name": "Quality C",
        "sector33": "機械",
        "close": 1500.0,
        "prev_close": 1490.0,
        "volume": 300_000,
        "trading_value": 450_000_000,
        "return_5d": 0.04,
        "return_20d": 0.12,
        "volume_ratio": 1.8,
        "ma20": 1450.0,
        "ma60": 1400.0,
        "ma20_deviation": 0.03,
        "price_date": "2026-07-10",
        "score": 80,
    },
    {
        "date": "2026-07-13",
        "rank": 4,
        "code": "4444",
        "name": "Quality D",
        "sector33": "化学",
        "close": None,
        "prev_close": 1000.0,
        "volume": 100_000,
        "trading_value": 100_000_000,
        "return_5d": None,
        "return_20d": None,
        "volume_ratio": None,
        "ma20": None,
        "ma60": None,
        "ma20_deviation": None,
        "price_date": "2026-07-13",
        "score": 75,
    },
])
original_scores = rows["score"].copy()
original_ranks = rows["rank"].copy()
graded = data_quality.attach_quality(
    rows,
    minimum_trading_value=100_000_000,
    policy_path=ROOT / data_quality.POLICY_PATH,
)
pd.testing.assert_series_equal(graded["score"], original_scores, check_names=False)
pd.testing.assert_series_equal(graded["rank"], original_ranks, check_names=False)
assert graded["data_quality_grade"].tolist() == ["A", "B", "C", "D"]
assert graded["data_quality_eligible_for_a"].tolist() == [True, True, False, False]
assert graded.loc[1, "data_quality_reason_codes"] == "BLANK_SECTOR"
assert "STALE_PRICE" in graded.loc[2, "data_quality_reason_codes"]
assert "INVALID_CLOSE" in graded.loc[3, "data_quality_reason_codes"]
assert graded.loc[2, "data_quality_current"] is False or graded.loc[2, "data_quality_current"] == False
assert graded.loc[3, "data_quality_core_complete"] is False or graded.loc[3, "data_quality_core_complete"] == False

corporate = rows.iloc[[0]].copy()
corporate.loc[corporate.index[0], "close"] = 1500.0
corporate.loc[corporate.index[0], "prev_close"] = 1000.0
corporate = data_quality.attach_quality(
    corporate,
    minimum_trading_value=100_000_000,
    policy_path=ROOT / data_quality.POLICY_PATH,
)
assert corporate.iloc[0]["data_quality_grade"] == "C"
assert corporate.iloc[0]["data_quality_corporate_action_suspected"] in {True, 1}
assert "POSSIBLE_CORPORATE_ACTION" in corporate.iloc[0]["data_quality_reason_codes"]

action_priority = pd.DataFrame([
    {"code": "1111", "name": "Quality A", "rank": 1, "action_priority": "A", "action_priority_score": 90},
    {"code": "2222", "name": "Quality B", "rank": 2, "action_priority": "A", "action_priority_score": 85},
    {"code": "3333", "name": "Quality C", "rank": 3, "action_priority": "A", "action_priority_score": 80},
    {"code": "4444", "name": "Quality D", "rank": 4, "action_priority": "A", "action_priority_score": 75},
])
gated = data_quality.apply_priority_gate(action_priority, graded)
by_code = gated.set_index("code")
assert by_code.loc["1111", "action_priority"] == "A"
assert by_code.loc["2222", "action_priority"] == "A"
assert by_code.loc["3333", "action_priority"] == "B"
assert by_code.loc["4444", "action_priority"] == "見送り"
assert by_code.loc["3333", "action_priority_before_quality"] == "A"
assert by_code.loc["4444", "action_priority_before_quality"] == "A"
assert not (
    gated["action_priority"].eq("A")
    & gated["data_quality_grade"].isin(["C", "D"])
).any()

mutable = action_priority.copy()
data_quality.replace_frame_in_place(mutable, gated)
pd.testing.assert_frame_equal(
    mutable.reset_index(drop=True),
    gated.reset_index(drop=True),
    check_dtype=False,
)

fields = data_quality.summary_fields(graded, gated)
assert fields["Data Quality評価件数"] == 4
assert fields["Data Quality A"] == 1
assert fields["Data Quality B"] == 1
assert fields["Data Quality C"] == 1
assert fields["Data Quality D"] == 1
assert fields["Data Quality優先度調整数"] == 2
assert fields["品質C/DのA候補"] == 0

plain = "\n".join(data_quality.plain_section(graded, gated))
assert "Top100: A 1 / B 1 / C 1 / D 1" in plain
assert "品質CはA昇格不可" in plain
assert "STALE" not in plain or "株価日付" in plain
html = data_quality.html_section(graded, gated)
assert "Data Quality" in html
assert "Momentumスコアと順位は不変" in html
assert "優先度調整 <b>2件</b>" in html

with TemporaryDirectory() as temporary:
    root = Path(temporary)
    history_path = root / "ranking.csv"
    main.write_ranking_history(graded, str(history_path))
    persisted = pd.read_csv(history_path, dtype={"code": str})
    assert set(data_quality.QUALITY_COLUMNS).issubset(persisted.columns)
    assert persisted["data_quality_grade"].tolist() == ["A", "B", "C", "D"]
    assert persisted["score"].tolist() == original_scores.tolist()
    assert persisted["rank"].tolist() == original_ranks.tolist()

    workbook_path = root / "report.xlsx"
    workbook = Workbook()
    workbook.active.title = "Summary"
    workbook.save(workbook_path)
    data_quality.patch_workbook(workbook_path, graded, gated)
    checked = load_workbook(workbook_path, data_only=True)
    assert "Data Quality" in checked.sheetnames
    sheet = checked["Data Quality"]
    values = [cell.value for row in sheet.iter_rows() for cell in row]
    assert "Grade A" in values
    assert "Priority adjustments" in values
    assert "NONE" in values
    assert "Quality C" in values
    assert "Quality D" in values

source = (ROOT / "daily_runner.py").read_text(encoding="utf-8")
assert "data_quality.attach_quality" in source
assert "data_quality.apply_priority_gate" in source
assert "data_quality.patch_workbook" in source
assert "main_module.attach_strategy_provenance = patched_provenance" in source
assert "main_module.excel_report = patched_excel" in source
assert "main.py" not in (ROOT / "research" / "data_quality_policy.yaml").read_text(encoding="utf-8")

print("data quality grading validation passed")
