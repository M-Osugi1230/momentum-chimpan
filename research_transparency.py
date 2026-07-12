"""Read-only research transparency helpers for the daily dashboard.

The canonical catalog governs the decision. A separately signed compact status
may add prospective progress counts and statistics after integrity validation.
Neither source can change rankings, score weights, thresholds, strategy
fingerprints, or production state. Invalid inputs fall back to a conservative hold.
"""
from __future__ import annotations

import html
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

import research_evidence_catalog as evidence_catalog
import volume_component_forward_status as forward_status

DEFAULT_CATALOG = "research/evidence_catalog.yaml"
DEFAULT_FORWARD_STATUS = "data/volume_component_forward_status.json"
SHEET_NAME = "Research Evidence"


def empty_progress() -> dict[str, dict[str, Any]]:
    return {
        str(horizon): {
            "horizon_days": horizon,
            "baseline_outcome_count": 0,
            "tested_outcome_count": 0,
            "minimum_variant_outcome_count": 0,
            "required_outcomes_per_variant": 100,
            "paired_date_count": 0,
            "required_paired_dates": 20,
            "outcome_progress_ratio": 0.0,
            "paired_date_progress_ratio": 0.0,
            "sample_adequate": False,
            "mean_daily_difference": None,
            "early_mean_difference": None,
            "late_mean_difference": None,
            "ci_low": None,
            "ci_high": None,
            "two_sided_p_value": None,
            "harm_p_value": None,
        }
        for horizon in (10, 20)
    }


def _fallback(error: str = "") -> dict[str, Any]:
    return {
        "catalog_health": "WARN",
        "catalog_error": error,
        "forward_status_health": "WARN",
        "forward_status_error": error,
        "forward_status_generated_at": "",
        "forward_status_fingerprint": "",
        "forward_status_source_run_id": "",
        "forward_progress": empty_progress(),
        "subject_id": "volume_ratio_score_component",
        "subject_label": "出来高倍率スコア構成要素",
        "production_weight_points": 15,
        "decision": evidence_catalog.HOLD_DECISION,
        "historical_consensus": "UNKNOWN",
        "research_status": "UNKNOWN",
        "governing_study_id": forward_status.STUDY_ID,
        "governing_study_status": "UNKNOWN",
        "next_decision_trigger": "FORWARD_EVIDENCE_GATE_COMPLETION",
        "automatic_weight_change_allowed": False,
        "automatic_strategy_change_allowed": False,
        "promotion_evidence_allowed": False,
        "manual_review_required": True,
        "decision_reason": "研究台帳を確認できないため、現行15点を維持して手動確認します。",
        "studies": [],
    }


def load_forward_progress(
    status_path: str | Path = DEFAULT_FORWARD_STATUS,
) -> tuple[dict[str, Any] | None, str]:
    try:
        payload = forward_status.load_json(status_path)
        errors = forward_status.validate_status(payload)
        if errors:
            return None, " / ".join(errors[:5])
        return payload, ""
    except Exception as exc:
        return None, str(exc)


def load_snapshot(
    catalog_path: str | Path = DEFAULT_CATALOG,
    repository_root: str | Path = ".",
    status_path: str | Path = DEFAULT_FORWARD_STATUS,
) -> dict[str, Any]:
    try:
        catalog = evidence_catalog.load_catalog(catalog_path)
        errors = evidence_catalog.validate_catalog(catalog, repository_root)
        if errors:
            return _fallback(" / ".join(errors[:5]))
        subject = catalog["subject"]
        studies = list(catalog.get("studies") or [])
        governing = next(
            (
                study
                for study in studies
                if study.get("id") == subject.get("governing_study_id")
            ),
            {},
        )
        signed_status, status_error = load_forward_progress(status_path)
        governing_status = governing.get("status", "UNKNOWN")
        progress = empty_progress()
        generated_at = ""
        fingerprint = ""
        source_run_id = ""
        status_health = "WARN"
        if signed_status is not None:
            if signed_status.get("study_id") != subject.get("governing_study_id"):
                status_error = "signed status study does not match governing catalog study"
            else:
                governing_status = signed_status.get("evidence_status", governing_status)
                progress = dict(signed_status.get("horizons") or empty_progress())
                generated_at = str(signed_status.get("generated_at_utc") or "")
                fingerprint = str(signed_status.get("evidence_fingerprint") or "")
                source_run_id = str(signed_status.get("source_run_id") or "")
                status_health = "PASS"
                status_error = ""
        return {
            "catalog_health": "PASS",
            "catalog_error": "",
            "forward_status_health": status_health,
            "forward_status_error": status_error,
            "forward_status_generated_at": generated_at,
            "forward_status_fingerprint": fingerprint,
            "forward_status_source_run_id": source_run_id,
            "forward_progress": progress,
            "subject_id": subject.get("id", ""),
            "subject_label": subject.get("label", ""),
            "production_weight_points": int(
                subject.get("current_production_weight_points", 15)
            ),
            "decision": subject.get("current_decision", evidence_catalog.HOLD_DECISION),
            "historical_consensus": subject.get("historical_consensus", "UNKNOWN"),
            "research_status": subject.get("current_research_status", "UNKNOWN"),
            "governing_study_id": subject.get("governing_study_id", ""),
            "governing_study_status": governing_status,
            "next_decision_trigger": subject.get("next_decision_trigger", ""),
            "automatic_weight_change_allowed": bool(
                subject.get("automatic_weight_change_allowed", False)
            ),
            "automatic_strategy_change_allowed": bool(
                subject.get("automatic_strategy_change_allowed", False)
            ),
            "promotion_evidence_allowed": bool(
                subject.get("promotion_evidence_allowed", False)
            ),
            "manual_review_required": bool(
                subject.get("manual_review_required", True)
            ),
            "decision_reason": subject.get("decision_reason", ""),
            "studies": studies,
        }
    except Exception as exc:
        return _fallback(str(exc))


def progress_text(record: dict[str, Any]) -> str:
    baseline = int(record.get("baseline_outcome_count", 0) or 0)
    tested = int(record.get("tested_outcome_count", 0) or 0)
    required = int(record.get("required_outcomes_per_variant", 100) or 100)
    paired = int(record.get("paired_date_count", 0) or 0)
    paired_required = int(record.get("required_paired_dates", 20) or 20)
    return (
        f"baseline {baseline}/{required}・除外 {tested}/{required}・"
        f"paired {paired}/{paired_required}日"
    )


def summary_fields(snapshot: dict[str, Any]) -> dict[str, Any]:
    progress = snapshot.get("forward_progress") or empty_progress()
    return {
        "研究台帳状態": snapshot.get("catalog_health", "WARN"),
        "Forward Status状態": snapshot.get("forward_status_health", "WARN"),
        "研究対象": snapshot.get("subject_label", "出来高倍率スコア構成要素"),
        "出来高倍率配点": snapshot.get("production_weight_points", 15),
        "研究判断": snapshot.get("decision", evidence_catalog.HOLD_DECISION),
        "歴史エビデンス": snapshot.get("historical_consensus", "UNKNOWN"),
        "研究ステータス": snapshot.get("research_status", "UNKNOWN"),
        "Forward Evidence": snapshot.get("governing_study_status", "UNKNOWN"),
        "Forward 10日進捗": progress_text(progress.get("10", {})),
        "Forward 20日進捗": progress_text(progress.get("20", {})),
        "Forward Status更新": snapshot.get("forward_status_generated_at", ""),
        "次回判断条件": snapshot.get("next_decision_trigger", ""),
        "自動配点変更": "DISABLED",
        "自動戦略変更": "DISABLED",
        "研究手動レビュー": "REQUIRED",
    }


def plain_section(snapshot: dict[str, Any]) -> list[str]:
    catalog_health = snapshot.get("catalog_health", "WARN")
    status_health = snapshot.get("forward_status_health", "WARN")
    progress = snapshot.get("forward_progress") or empty_progress()
    lines = [
        "【研究エビデンスの現在地】",
        (
            f"対象: {snapshot.get('subject_label', '出来高倍率スコア構成要素')} / "
            f"現行配点 {snapshot.get('production_weight_points', 15)}点（据え置き）"
        ),
        (
            f"判断: {snapshot.get('decision', evidence_catalog.HOLD_DECISION)} / "
            f"研究状態: {snapshot.get('research_status', 'UNKNOWN')}"
        ),
        (
            f"歴史検証: {snapshot.get('historical_consensus', 'UNKNOWN')} / "
            f"Forward Evidence: {snapshot.get('governing_study_status', 'UNKNOWN')}"
        ),
        f"Forward 10日: {progress_text(progress.get('10', {}))}",
        f"Forward 20日: {progress_text(progress.get('20', {}))}",
        "自動配点変更・自動戦略変更は無効です。次の判断はforward evidence完了後の手動レビューです。",
    ]
    if catalog_health != "PASS":
        lines.append(
            "注意: 研究台帳を完全検証できないため、安全側で現行配点を維持しています。"
        )
    elif status_health != "PASS":
        lines.append(
            "注意: Forward statusを検証できないため、件数は安全側の0表示とし、現行配点を維持しています。"
        )
    lines.append("")
    return lines


def html_section(snapshot: dict[str, Any]) -> str:
    catalog_health = snapshot.get("catalog_health", "WARN")
    status_health = snapshot.get("forward_status_health", "WARN")
    healthy = catalog_health == "PASS" and status_health == "PASS"
    border = "#7c3aed" if healthy else "#b45309"
    background = "#faf5ff" if healthy else "#fffbeb"
    label = (
        f"{snapshot.get('research_status', 'UNRESOLVED')} / "
        f"FORWARD {snapshot.get('governing_study_status', 'UNKNOWN')}"
    )
    if not healthy:
        label = "SAFE HOLD / STATUS WARN"
    esc = lambda value: html.escape(str(value or ""))
    progress = snapshot.get("forward_progress") or empty_progress()
    warning = ""
    if not healthy:
        warning = (
            '<div style="font-size:11px;color:#b45309;margin-top:6px">'
            "台帳または署名statusを完全検証できないため、安全側で現行配点を維持しています。</div>"
        )
    return f'''<div style="background:{background};border:2px solid {border};border-radius:18px;padding:16px;margin-top:14px">
<div style="font-size:12px;font-weight:900;color:{border}">RESEARCH EVIDENCE</div>
<div style="font-size:19px;font-weight:900;color:#3b0764;margin-top:2px">{esc(snapshot.get("subject_label", "出来高倍率スコア構成要素"))} <span style="float:right;font-size:12px;color:{border}">{esc(label)}</span></div>
<div style="clear:both;font-size:13px;color:#334155;margin-top:8px">現行配点 <b>{int(snapshot.get("production_weight_points", 15))}点のまま維持</b> ・ 自動配点変更なし ・ 自動戦略変更なし</div>
<div style="font-size:12px;line-height:1.8;color:#475569;margin-top:5px">歴史検証 <b>{esc(snapshot.get("historical_consensus", "UNKNOWN"))}</b> ・ Forward Evidence <b>{esc(snapshot.get("governing_study_status", "UNKNOWN"))}</b><br>10日: {esc(progress_text(progress.get("10", {})))}<br>20日: {esc(progress_text(progress.get("20", {})))}<br>次の判断はforward evidence完了後の手動レビューです。</div>{warning}</div>'''


def evidence_rows(snapshot: dict[str, Any]) -> pd.DataFrame:
    rows: list[dict[str, Any]] = [
        {
            "section": "Current Decision",
            "study_id": snapshot.get("subject_id", ""),
            "label": snapshot.get("subject_label", ""),
            "evidence_class": "GOVERNED_DECISION",
            "status": snapshot.get("research_status", "UNKNOWN"),
            "decision": snapshot.get("decision", evidence_catalog.HOLD_DECISION),
            "weight_points": snapshot.get("production_weight_points", 15),
            "source_pr": None,
            "delta_excess_return": None,
            "two_sided_p_value": None,
            "interpretation": snapshot.get("decision_reason", ""),
        }
    ]
    for study in snapshot.get("studies", []):
        rows.append(
            {
                "section": "Study",
                "study_id": study.get("id", ""),
                "label": study.get("label", ""),
                "evidence_class": study.get("evidence_class", ""),
                "status": study.get("status", ""),
                "decision": "",
                "weight_points": None,
                "source_pr": study.get("source_pr"),
                "delta_excess_return": study.get("primary_delta_excess_return"),
                "two_sided_p_value": study.get("two_sided_p_value"),
                "interpretation": study.get("interpretation", ""),
            }
        )
    progress = snapshot.get("forward_progress") or empty_progress()
    for horizon in (10, 20):
        record = progress.get(str(horizon), {})
        ci_low = record.get("ci_low")
        ci_high = record.get("ci_high")
        ci_text = "CI未算定" if ci_low is None or ci_high is None else f"CI {ci_low:+.6f}～{ci_high:+.6f}"
        rows.append(
            {
                "section": "Forward Progress",
                "study_id": f"{snapshot.get('governing_study_id', forward_status.STUDY_ID)}:{horizon}d",
                "label": f"{horizon}営業日forward",
                "evidence_class": "PROSPECTIVE_PROGRESS",
                "status": snapshot.get("governing_study_status", "UNKNOWN"),
                "decision": progress_text(record),
                "weight_points": None,
                "source_pr": 61,
                "delta_excess_return": record.get("mean_daily_difference"),
                "two_sided_p_value": record.get("two_sided_p_value"),
                "interpretation": f"{ci_text} / status updated {snapshot.get('forward_status_generated_at', '')}",
            }
        )
    return pd.DataFrame(rows)


def patch_workbook(path: str | Path, snapshot: dict[str, Any]) -> None:
    target = Path(path)
    if not target.exists():
        raise FileNotFoundError(target)
    workbook = load_workbook(target)
    if SHEET_NAME in workbook.sheetnames:
        del workbook[SHEET_NAME]
    sheet = workbook.create_sheet(SHEET_NAME, 1)
    frame = evidence_rows(snapshot)
    columns = list(frame.columns)
    for column_index, value in enumerate(columns, start=1):
        cell = sheet.cell(row=1, column=column_index, value=value)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="6D28D9")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for row_index, row in enumerate(frame.itertuples(index=False), start=2):
        for column_index, value in enumerate(row, start=1):
            sheet.cell(row=row_index, column=column_index, value=value)
    sheet.freeze_panes = "A2"
    widths = {
        "A": 20,
        "B": 48,
        "C": 34,
        "D": 34,
        "E": 28,
        "F": 60,
        "G": 14,
        "H": 12,
        "I": 20,
        "J": 20,
        "K": 88,
    }
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    workbook.save(target)
