"""Validate whether one Daily Momentum Report run is eligible for live evidence.

The validator reads only the exact upstream workflow artifact. It combines the
production, evidence, recovery, data-quality, Daily Research Focus, SMTP, and
prospective-outcome prerequisites into one signed readiness result. It cannot
write repository state or change strategy behavior.
"""
from __future__ import annotations

import argparse
import hashlib
import json
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook

import email_delivery
import priority_outcomes

READINESS_VERSION = "2026-07-13-live-session-readiness-v1"
ELIGIBLE_DATE_FROM = "2026-07-13"
DEFAULT_OUTPUT_DIR = "output/live-session-readiness"
REQUIRED_WORKBOOK_SHEETS = {
    "Summary",
    "Momentum Top100",
    "Action Priority",
    "Daily Action List",
    "Data Quality",
    "Research Evidence",
}
EXPLANATION_COLUMNS = {
    "why_today",
    "what_changed",
    "risk_summary",
    "next_research_questions",
}
QUALITY_COLUMNS = {
    "data_quality_grade",
    "data_quality_score",
    "data_quality_eligible_for_a",
    "data_quality_reason_codes",
    "data_quality_warnings",
}


def canonical_hash(payload: Any) -> str:
    text = json.dumps(payload, ensure_ascii=False, sort_keys=True, separators=(",", ":"))
    return hashlib.sha256(text.encode("utf-8")).hexdigest()


def sha256_file(path: Path) -> str:
    if not path.is_file():
        return ""
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def valid_sha256(value: Any) -> bool:
    text = str(value or "").strip().lower()
    return len(text) == 64 and all(character in "0123456789abcdef" for character in text)


def optional_text(value: Any) -> str:
    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except (TypeError, ValueError):
        pass
    text = str(value).strip()
    return "" if text.lower() in {"", "nan", "none", "nat"} else text


def to_bool(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)) and not pd.isna(value):
        return bool(value)
    return optional_text(value).lower() in {"true", "1", "yes", "y"}


def to_float(value: Any) -> float | None:
    converted = pd.to_numeric(pd.Series([value]), errors="coerce").iloc[0]
    return None if pd.isna(converted) else float(converted)


def find_file(root: Path, name: str) -> Path | None:
    matches = sorted(path for path in root.rglob(name) if path.is_file()) if root.exists() else []
    return matches[0] if matches else None


def load_json(path: Path | None) -> dict[str, Any]:
    if path is None or not path.is_file():
        return {}
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return {}
    return payload if isinstance(payload, dict) else {}


def load_csv(path: Path | None, dtype: dict[str, Any] | None = None) -> pd.DataFrame:
    if path is None or not path.is_file() or path.stat().st_size == 0:
        return pd.DataFrame()
    try:
        return pd.read_csv(path, dtype=dtype)
    except Exception:
        return pd.DataFrame()


def workbook_sheet_names(path: Path | None) -> set[str]:
    if path is None or not path.is_file():
        return set()
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except Exception:
        return set()
    names = set(workbook.sheetnames)
    workbook.close()
    return names


def read_sheet(path: Path | None, sheet_name: str, dtype: dict[str, Any] | None = None) -> pd.DataFrame:
    if path is None or not path.is_file():
        return pd.DataFrame()
    try:
        return pd.read_excel(path, sheet_name=sheet_name, dtype=dtype)
    except Exception:
        return pd.DataFrame()


def artifact_fingerprint(root: Path) -> str:
    entries = [
        {
            "path": path.relative_to(root).as_posix(),
            "size": path.stat().st_size,
            "sha256": sha256_file(path),
        }
        for path in sorted(candidate for candidate in root.rglob("*") if candidate.is_file())
    ] if root.exists() else []
    return canonical_hash(entries)


def gate(name: str, status: str, detail: str, metrics: dict[str, Any] | None = None) -> dict[str, Any]:
    return {
        "name": name,
        "status": status,
        "detail": detail,
        "metrics": metrics or {},
    }


def build_readiness(
    artifact_root: str | Path,
    *,
    source_run_id: str,
    source_run_url: str,
    upstream_conclusion: str,
    upstream_event: str,
    head_sha: str,
    created_at_utc: str,
    updated_at_utc: str,
    generated_at_utc: str | None = None,
) -> dict[str, Any]:
    root = Path(artifact_root)
    report_path = find_file(root, "daily_report.xlsx")
    heartbeat = load_json(find_file(root, "operations_heartbeat.json"))
    fingerprint_manifest = load_json(find_file(root, "strategy_fingerprint.json"))
    evidence = load_json(find_file(root, "evidence_stamp_audit.json"))
    recovery = load_json(find_file(root, "recovery_snapshot_audit.json"))
    maintenance = load_json(find_file(root, "state_maintenance.json"))
    email_receipt = load_json(find_file(root, "email_delivery_receipt.json"))
    ranking = load_csv(find_file(root, "momentum_daily_ranking.csv"), dtype={"code": str})
    sheet_names = workbook_sheet_names(report_path)
    top100 = read_sheet(report_path, "Momentum Top100", dtype={"code": str})
    action = read_sheet(report_path, "Action Priority", dtype={"code": str})
    summary = read_sheet(report_path, "Summary")

    report_date = optional_text(heartbeat.get("report_date"))
    if not report_date and not summary.empty:
        report_date = optional_text(summary.iloc[0].get("実行日"))
    strategy_fingerprint = optional_text(fingerprint_manifest.get("strategy_fingerprint"))
    state_update = heartbeat.get("state_update_executed") is True
    gates: list[dict[str, Any]] = []

    upstream_ok = str(upstream_conclusion).lower() == "success"
    gates.append(gate(
        "upstream_workflow",
        "PASS" if upstream_ok else "FAIL",
        f"upstream conclusion={upstream_conclusion}",
    ))

    required_files = {
        "daily_report.xlsx": report_path,
        "operations_heartbeat.json": find_file(root, "operations_heartbeat.json"),
        "strategy_fingerprint.json": find_file(root, "strategy_fingerprint.json"),
        "momentum_daily_ranking.csv": find_file(root, "momentum_daily_ranking.csv"),
        "evidence_stamp_audit.json": find_file(root, "evidence_stamp_audit.json"),
        "recovery_snapshot_audit.json": find_file(root, "recovery_snapshot_audit.json"),
        "state_maintenance.json": find_file(root, "state_maintenance.json"),
        "email_delivery_receipt.json": find_file(root, "email_delivery_receipt.json"),
    }
    missing_files = sorted(name for name, path in required_files.items() if path is None)
    gates.append(gate(
        "artifact_completeness",
        "PASS" if not missing_files else "FAIL",
        "all required files present" if not missing_files else "missing: " + ", ".join(missing_files),
        {"required_file_count": len(required_files), "missing_file_count": len(missing_files)},
    ))

    heartbeat_ok = bool(
        heartbeat
        and str(heartbeat.get("workflow_status", "")).upper() == "SUCCESS"
        and state_update
        and report_date >= ELIGIBLE_DATE_FROM
    )
    gates.append(gate(
        "production_heartbeat",
        "PASS" if heartbeat_ok else "FAIL",
        "successful full state update on an eligible date" if heartbeat_ok else "heartbeat/status/date is not eligible",
        {
            "report_date": report_date,
            "eligible_date_from": ELIGIBLE_DATE_FROM,
            "state_update_executed": state_update,
            "workflow_status": heartbeat.get("workflow_status"),
            "current_day_price_ratio": heartbeat.get("current_day_price_ratio"),
        },
    ))

    fingerprint_ok = valid_sha256(strategy_fingerprint)
    gates.append(gate(
        "strategy_fingerprint",
        "PASS" if fingerprint_ok else "FAIL",
        "non-empty governed SHA-256 strategy fingerprint" if fingerprint_ok else "strategy fingerprint missing or invalid",
        {"strategy_fingerprint": strategy_fingerprint},
    ))

    missing_sheets = sorted(REQUIRED_WORKBOOK_SHEETS - sheet_names)
    gates.append(gate(
        "workbook_contract",
        "PASS" if not missing_sheets else "FAIL",
        "all required daily sheets present" if not missing_sheets else "missing sheets: " + ", ".join(missing_sheets),
        {"sheet_count": len(sheet_names), "missing_sheet_count": len(missing_sheets)},
    ))

    ranking_rows = pd.DataFrame()
    ranking_duplicates = None
    ranking_fingerprint_match = False
    if not ranking.empty and {"date", "code"}.issubset(ranking.columns):
        ranking_rows = ranking[ranking["date"].astype(str).eq(report_date)].copy()
        ranking_duplicates = int(ranking.duplicated(["date", "code"], keep=False).sum())
        if (
            not ranking_rows.empty
            and fingerprint_ok
            and "strategy_fingerprint" in ranking_rows.columns
        ):
            ranking_fingerprint_match = bool(
                ranking_rows["strategy_fingerprint"]
                .fillna("")
                .astype(str)
                .str.strip()
                .eq(strategy_fingerprint)
                .all()
            )
    ranking_ok = bool(
        not ranking_rows.empty
        and ranking_duplicates == 0
        and ranking_fingerprint_match
    )
    gates.append(gate(
        "ranking_history",
        "PASS" if ranking_ok else "FAIL",
        "eligible ranking rows are unique and fingerprint-consistent" if ranking_ok else "ranking rows, duplicates, or fingerprint are invalid",
        {
            "report_date_row_count": len(ranking_rows),
            "duplicate_row_count": ranking_duplicates,
            "fingerprint_matches": ranking_fingerprint_match,
        },
    ))

    evidence_ok = bool(
        evidence
        and optional_text(evidence.get("strategy_fingerprint")) == strategy_fingerprint
        and int(evidence.get("stamped_rows", 0) or 0) > 0
    )
    gates.append(gate(
        "evidence_stamp",
        "PASS" if evidence_ok else "FAIL",
        "ranking/report evidence is stamped with the governed fingerprint" if evidence_ok else "evidence stamp is missing, empty, or mismatched",
        {
            "stamped_rows": int(evidence.get("stamped_rows", 0) or 0),
            "fingerprint_matches": optional_text(evidence.get("strategy_fingerprint")) == strategy_fingerprint,
        },
    ))

    recovery_ok = bool(
        recovery
        and recovery.get("status") == "SEALED"
        and recovery.get("complete") is True
    )
    gates.append(gate(
        "recovery_snapshot",
        "PASS" if recovery_ok else "FAIL",
        "recoverable state snapshot is sealed and complete" if recovery_ok else "recovery snapshot is not sealed and complete",
        {"status": recovery.get("status"), "complete": recovery.get("complete")},
    ))

    maintenance_ok = bool(
        maintenance
        and maintenance.get("validation_status") == "PASS"
        and int(maintenance.get("validation_failures", 0) or 0) == 0
    )
    gates.append(gate(
        "state_maintenance",
        "PASS" if maintenance_ok else "FAIL",
        "state maintenance validation passed" if maintenance_ok else "state maintenance validation failed or is missing",
        {
            "validation_status": maintenance.get("validation_status"),
            "validation_failures": int(maintenance.get("validation_failures", 0) or 0),
        },
    ))

    top100_quality_columns = QUALITY_COLUMNS.issubset(top100.columns)
    grades = top100.get("data_quality_grade", pd.Series(index=top100.index, dtype=str)).fillna("").astype(str)
    quality_coverage = float(grades.isin(["A", "B", "C", "D"]).mean()) if len(top100) else 0.0
    quality_ok = bool(not top100.empty and top100_quality_columns and quality_coverage == 1.0)
    gates.append(gate(
        "data_quality",
        "PASS" if quality_ok else "FAIL",
        "every Top100 row has a valid A/B/C/D quality grade" if quality_ok else "Top100 quality columns or coverage are incomplete",
        {
            "top100_count": len(top100),
            "quality_coverage": quality_coverage,
            "grade_a": int((grades == "A").sum()),
            "grade_b": int((grades == "B").sum()),
            "grade_c": int((grades == "C").sum()),
            "grade_d": int((grades == "D").sum()),
        },
    ))

    action_required_columns = {
        "research_bucket",
        "daily_action_list",
        "data_quality_grade",
        *EXPLANATION_COLUMNS,
    }
    action_columns_ok = action_required_columns.issubset(action.columns)
    bucket = action.get("research_bucket", pd.Series(index=action.index, dtype=str)).fillna("").astype(str)
    action_list_mask = action.get("daily_action_list", pd.Series(False, index=action.index)).map(to_bool)
    a_count = int(bucket.eq("A").sum())
    action_list_count = int(action_list_mask.sum())
    invalid_a = int(
        (
            bucket.eq("A")
            & action.get("data_quality_grade", pd.Series(index=action.index, dtype=str)).fillna("").astype(str).isin(["C", "D"])
        ).sum()
    )
    explanation_complete = pd.Series(True, index=action.index, dtype=bool)
    for column in EXPLANATION_COLUMNS:
        if column not in action.columns:
            explanation_complete &= False
        else:
            explanation_complete &= action[column].fillna("").astype(str).str.strip().ne("")
    focus_ok = bool(
        not action.empty
        and action_columns_ok
        and a_count <= 5
        and action_list_count <= 10
        and invalid_a == 0
        and explanation_complete.all()
    )
    gates.append(gate(
        "daily_research_focus",
        "PASS" if focus_ok else "FAIL",
        "A cap, action-list cap, quality boundary, and explanations pass" if focus_ok else "Daily Research Focus contract failed",
        {
            "candidate_count": len(action),
            "priority_a_count": a_count,
            "daily_action_list_count": action_list_count,
            "quality_c_or_d_in_a_count": invalid_a,
            "explanation_complete_count": int(explanation_complete.sum()),
        },
    ))

    receipt_issues = email_delivery.validate_receipt(email_receipt) if email_receipt else ["receipt missing"]
    receipt_status = optional_text(email_receipt.get("status"))
    if receipt_issues:
        email_gate_status = "FAIL"
        email_detail = "; ".join(receipt_issues)
    elif receipt_status == "SMTP_ACCEPTED":
        email_gate_status = "PASS"
        email_detail = "configured SMTP server accepted the daily email"
    elif receipt_status == "SKIPPED_SECRETS_MISSING":
        email_gate_status = "REVIEW_REQUIRED"
        email_detail = "daily email was skipped because required secrets were missing"
    else:
        email_gate_status = "FAIL"
        email_detail = f"email receipt status={receipt_status}"
    gates.append(gate(
        "smtp_acceptance",
        email_gate_status,
        email_detail,
        {
            "receipt_status": receipt_status,
            "recipient_count": int(email_receipt.get("recipient_count", 0) or 0),
            "inbox_delivery_claimed": email_receipt.get("inbox_delivery_claimed"),
        },
    ))

    priority_policy = priority_outcomes.load_policy()
    extractable_decisions = priority_outcomes.extract_decisions(
        root,
        source_run_id=source_run_id,
        source_run_url=source_run_url,
        recorded_at_utc=updated_at_utc or generated_at_utc or datetime.now(timezone.utc).isoformat(timespec="seconds"),
        policy=priority_policy,
    )
    outcome_ready = bool(not extractable_decisions.empty)
    gates.append(gate(
        "priority_outcome_ingestion",
        "PASS" if outcome_ready else "FAIL",
        "daily decisions can be ingested into the prospective 5/10/20-session tracker" if outcome_ready else "no eligible priority decisions can be extracted",
        {
            "extractable_decision_count": len(extractable_decisions),
            "daily_action_decision_count": int(extractable_decisions.get("daily_action_list", pd.Series(dtype=bool)).sum()) if not extractable_decisions.empty else 0,
        },
    ))

    forward_ready = bool(
        report_date >= ELIGIBLE_DATE_FROM
        and fingerprint_ok
        and not ranking_rows.empty
        and ranking_fingerprint_match
    )
    gates.append(gate(
        "forward_evidence_prerequisites",
        "PASS" if forward_ready else "FAIL",
        "ranking date and strategy stamp are eligible for Forward Evidence" if forward_ready else "Forward Evidence prerequisites are incomplete",
        {
            "eligible_signal_date_from": ELIGIBLE_DATE_FROM,
            "report_date": report_date,
            "eligible_ranking_rows": len(ranking_rows),
        },
    ))

    freshness_ratio = to_float(heartbeat.get("current_day_price_ratio"))
    if freshness_ratio is None:
        freshness_status = "REVIEW_REQUIRED"
        freshness_detail = "current-day price ratio is unavailable"
    elif freshness_ratio >= 0.98:
        freshness_status = "PASS"
        freshness_detail = "current-day price coverage meets the 98% operating gate"
    else:
        freshness_status = "REVIEW_REQUIRED"
        freshness_detail = "current-day price coverage is below 98%"
    gates.append(gate(
        "market_data_coverage",
        freshness_status,
        freshness_detail,
        {"current_day_price_ratio": freshness_ratio, "minimum_required": 0.98},
    ))

    fail_count = sum(item["status"] == "FAIL" for item in gates)
    review_count = sum(item["status"] == "REVIEW_REQUIRED" for item in gates)
    state = "FAIL" if fail_count else "REVIEW_REQUIRED" if review_count else "PASS"
    substantive = {
        "readiness_version": READINESS_VERSION,
        "readiness_state": state,
        "source_run_id": str(source_run_id),
        "source_run_url": source_run_url,
        "upstream_conclusion": upstream_conclusion,
        "upstream_event": upstream_event,
        "head_sha": head_sha,
        "created_at_utc": created_at_utc,
        "updated_at_utc": updated_at_utc,
        "report_date": report_date,
        "eligible_date_from": ELIGIBLE_DATE_FROM,
        "strategy_fingerprint": strategy_fingerprint,
        "artifact_fingerprint": artifact_fingerprint(root),
        "gate_count": len(gates),
        "pass_count": sum(item["status"] == "PASS" for item in gates),
        "review_required_count": review_count,
        "fail_count": fail_count,
        "gates": gates,
        "eligible_for_forward_evidence": forward_ready and state != "FAIL",
        "eligible_for_priority_outcome_ingestion": outcome_ready and state != "FAIL",
        "smtp_acceptance_only": True,
        "inbox_delivery_claimed": False,
        "automatic_score_change": False,
        "automatic_weight_change": False,
        "automatic_strategy_change": False,
        "automatic_priority_rule_change": False,
        "production_state_mutations": [],
        "research_only": True,
    }
    payload = {
        **substantive,
        "generated_at_utc": generated_at_utc or datetime.now(timezone.utc).isoformat(timespec="seconds"),
        "readiness_fingerprint": canonical_hash(substantive),
    }
    payload["status_sha256"] = canonical_hash(payload)
    return payload


def validate_readiness(payload: dict[str, Any]) -> list[str]:
    issues: list[str] = []
    if payload.get("readiness_version") != READINESS_VERSION:
        issues.append("invalid readiness_version")
    if payload.get("readiness_state") not in {"PASS", "REVIEW_REQUIRED", "FAIL"}:
        issues.append("invalid readiness_state")
    gates = payload.get("gates")
    if not isinstance(gates, list) or len(gates) != int(payload.get("gate_count", -1)):
        issues.append("gate count mismatch")
    else:
        valid_statuses = {"PASS", "REVIEW_REQUIRED", "FAIL"}
        if any(item.get("status") not in valid_statuses for item in gates if isinstance(item, dict)):
            issues.append("invalid gate status")
        fail_count = sum(item.get("status") == "FAIL" for item in gates if isinstance(item, dict))
        review_count = sum(item.get("status") == "REVIEW_REQUIRED" for item in gates if isinstance(item, dict))
        expected_state = "FAIL" if fail_count else "REVIEW_REQUIRED" if review_count else "PASS"
        if payload.get("readiness_state") != expected_state:
            issues.append("readiness state does not match gate statuses")
    if payload.get("smtp_acceptance_only") is not True:
        issues.append("smtp_acceptance_only must be true")
    if payload.get("inbox_delivery_claimed") is not False:
        issues.append("inbox delivery must not be claimed")
    for key in (
        "automatic_score_change",
        "automatic_weight_change",
        "automatic_strategy_change",
        "automatic_priority_rule_change",
    ):
        if payload.get(key) is not False:
            issues.append(f"{key} must be false")
    if payload.get("production_state_mutations") != []:
        issues.append("production_state_mutations must be empty")
    status_copy = dict(payload)
    supplied_status_hash = status_copy.pop("status_sha256", "")
    if supplied_status_hash != canonical_hash(status_copy):
        issues.append("status_sha256 mismatch")
    substantive = dict(status_copy)
    substantive.pop("generated_at_utc", None)
    supplied_fingerprint = substantive.pop("readiness_fingerprint", "")
    if supplied_fingerprint != canonical_hash(substantive):
        issues.append("readiness_fingerprint mismatch")
    return issues


def markdown_report(payload: dict[str, Any]) -> str:
    lines = [
        "# Live Session Readiness",
        "",
        f"State: **{payload['readiness_state']}**",
        f"Report date: **{payload.get('report_date') or '-'}**",
        f"Source run: `{payload.get('source_run_id') or '-'}`",
        f"Strategy fingerprint: `{payload.get('strategy_fingerprint') or '-'}`",
        "",
        "| Gate | Status | Detail |",
        "|---|---|---|",
    ]
    for item in payload.get("gates", []):
        lines.append(f"| {item['name']} | {item['status']} | {item['detail']} |")
    lines.extend([
        "",
        "## Evidence eligibility",
        "",
        f"- Forward Evidence eligible: **{payload['eligible_for_forward_evidence']}**",
        f"- Priority outcome ingestion eligible: **{payload['eligible_for_priority_outcome_ingestion']}**",
        "- Email claim: **SMTP acceptance only**",
        "- Inbox delivery claimed: **False**",
        "",
        "## Governance",
        "",
        "This validator is read-only. It does not change production state, scores, weights, strategy, priority rules, paper execution, or live orders.",
        "",
    ])
    return "\n".join(lines)


def write_outputs(payload: dict[str, Any], output_dir: str | Path) -> dict[str, str]:
    target = Path(output_dir)
    target.mkdir(parents=True, exist_ok=True)
    json_path = target / "live_session_readiness.json"
    markdown_path = target / "live_session_readiness.md"
    json_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    markdown_path.write_text(markdown_report(payload), encoding="utf-8")
    return {"json": str(json_path), "markdown": str(markdown_path)}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Validate exact-artifact live session readiness")
    subparsers = parser.add_subparsers(dest="command", required=True)
    build = subparsers.add_parser("build")
    build.add_argument("--artifact-root", required=True)
    build.add_argument("--source-run-id", required=True)
    build.add_argument("--source-run-url", default="")
    build.add_argument("--upstream-conclusion", required=True)
    build.add_argument("--upstream-event", required=True)
    build.add_argument("--head-sha", default="")
    build.add_argument("--created-at-utc", default="")
    build.add_argument("--updated-at-utc", default="")
    build.add_argument("--output-dir", default=DEFAULT_OUTPUT_DIR)
    build.add_argument("--generated-at-utc", default="")
    validate = subparsers.add_parser("validate")
    validate.add_argument("--json", required=True)
    return parser.parse_args()


def main_cli() -> int:
    args = parse_args()
    if args.command == "build":
        payload = build_readiness(
            args.artifact_root,
            source_run_id=args.source_run_id,
            source_run_url=args.source_run_url,
            upstream_conclusion=args.upstream_conclusion,
            upstream_event=args.upstream_event,
            head_sha=args.head_sha,
            created_at_utc=args.created_at_utc,
            updated_at_utc=args.updated_at_utc,
            generated_at_utc=args.generated_at_utc or None,
        )
        issues = validate_readiness(payload)
        if issues:
            print(json.dumps({"valid": False, "issues": issues}, ensure_ascii=False, indent=2))
            return 1
        outputs = write_outputs(payload, args.output_dir)
        print(json.dumps({"readiness": payload, "outputs": outputs}, ensure_ascii=False, indent=2))
        return 1 if payload["readiness_state"] == "FAIL" else 0
    payload = json.loads(Path(args.json).read_text(encoding="utf-8"))
    issues = validate_readiness(payload)
    print(json.dumps({"valid": not issues, "issues": issues}, ensure_ascii=False, indent=2))
    return 1 if issues else 0


if __name__ == "__main__":
    raise SystemExit(main_cli())
